"""
clean_tracker.py — STAGE 1 of the pipeline: clean the NC tracker BEFORE it goes
into the database. Read-only on the source; writes a cleaned copy + a review
report. Nothing here touches quality.db.

Why this exists (Adriele, 11.08.2026): the tracker is typed by several people
and Project / Flight Unit are often blank, misspelled, or put in the wrong
column. We do NOT want the database to inherit that mess. The tracker's own
`Set_Up` sheet is the ground truth — it lists every valid Project and, per
project, every valid Flight Unit. This engine normalises against that list,
recovers a Project from an unambiguous Flight Unit, blanks values that are not
valid, and FLAGS anything it cannot resolve so a human fixes it in the tracker.

This is a data-contract layer: it decides field content, so the rules are
written out explicitly rather than guessed. It never invents a value — when it
is not sure, it blanks the field and flags the row.

Scope of THIS file: Project + Flight Unit only (the explicit ask). Validating
Detection / Failure / Owner / Classification against Set_Up is left as a future
stage — registered here, not implemented.

Run:  cd ~/bgtools/dash && ./quality/bin/python cleaning/clean_tracker.py
Out:  cleaning/cleaned_tracker.xlsx   (sheets: clean · flags · summary)
"""
import os
import re
import sys

import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(os.path.dirname(HERE), "data")
TRACKER_SHEET = "NC_Tracker_Black_Out"
SETUP_SHEET = "Set_Up"
OUT = os.path.join(HERE, "cleaned_tracker.xlsx")

# The 8 canonical projects the dashboard understands. Flexline PLF / ISA are one
# project 'Flexline' at this level; the Flight Unit keeps the sub-detail.
CANON_PROJECTS = ["Ariane", "Vega", "Vega-C", "MHI_H3", "Relativity",
                  "SAS", "Vulcan", "Flexline"]

# Explicit alias map for the messy Project values actually seen in the tracker.
# Left = uppercased raw (trimmed); right = canonical. Extend as new spellings
# appear. Everything not covered here falls to the keyword rules, then to
# flight-unit inference, then to a flag.
PROJECT_ALIASES = {
    "ARIANE": "Ariane", "ARIANE 6": "Ariane", "ARIANE6": "Ariane",
    "ARIANE C6010S": "Ariane",
    "VEGA": "Vega", "VEGA-C": "Vega-C", "VEGA C": "Vega-C",
    "RELATIVITY": "Relativity", "RELATIVITY SPACE FM03": "Relativity",
    "SAS": "SAS",
    "VULCAN": "Vulcan", "VCN": "Vulcan", "VULCAN FRAME LAYUPS": "Vulcan",
    "MHI_H3": "MHI_H3", "MHI H3": "MHI_H3", "H3 INTEGRATION T-HALF": "MHI_H3",
    "FLEXLINE": "Flexline", "FLEXLINE PLF": "Flexline",
    "FLEXLINE ISA 1-2": "Flexline", "FLEXLINE ISA 2-3": "Flexline",
    "FLEXLINE FM05 / ISA 2-3": "Flexline", "FLEXLINE ISA14 1-2": "Flexline",
    "FLEXLINE FM4+": "Flexline",
}

# Values that explicitly mean 'no project' — blank them, do not flag as an error.
PROJECT_BLANKS = {"", "TBD", "ALL", "N/A", "-", "NONE", "??", "?"}


def _s(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def load_setup_vocab(path):
    """Build the ground-truth vocab from the Set_Up sheet:
    canonical projects, and Flight-Unit -> {projects} (a unit can be valid for
    more than one project — the bare FM codes are shared)."""
    su = pd.read_excel(path, sheet_name=SETUP_SHEET, header=None, dtype=str)
    header = [_s(x) for x in su.iloc[0, 1:8].tolist()]   # Ariane..Flexline cols
    col_proj = {"Ariane": "Ariane", "Vega": "Vega", "MHI H3": "MHI_H3",
                "Relativity": "Relativity", "SAS": "SAS", "Vulcan": "Vulcan",
                "Flexline": "Flexline"}
    fu_to_projects = {}
    for ci, name in enumerate(header, start=1):
        proj = col_proj.get(name, name)
        for v in su.iloc[1:, ci].dropna():
            code = _s(v)
            if code:
                fu_to_projects.setdefault(code.upper(), set()).add(proj)
    return fu_to_projects


def canon_project(raw):
    """Raw Project text -> a canonical project, or '' if it means nothing.
    Returns (value, matched_how)."""
    s = _s(raw)
    up = s.upper()
    if up in PROJECT_BLANKS:
        return "", "blank"
    if up in PROJECT_ALIASES:
        return PROJECT_ALIASES[up], "alias"
    # keyword fallback (specific before generic)
    if "VEGA-C" in up or "VEGA C" in up:
        return "Vega-C", "keyword"
    if "VEGA" in up:
        return "Vega", "keyword"
    if "FLEXLINE" in up or "FLEX LINE" in up:
        return "Flexline", "keyword"
    if "VULCAN" in up or up.startswith("VCN"):
        return "Vulcan", "keyword"
    if "RELATIVITY" in up:
        return "Relativity", "keyword"
    if up.startswith("MHI") or "H3" in re.split(r"[ _/()-]+", up):
        return "MHI_H3", "keyword"
    if "ARIANE" in up or re.match(r"A6\d", up) or re.match(r"C6\d", up):
        return "Ariane", "keyword"
    if up == "SAS" or "SAS" in re.split(r"[ _/()-]+", up):
        return "SAS", "keyword"
    return "", "unresolved"


def canon_flight_unit(raw, fu_vocab):
    """Flight Unit -> a valid Set_Up code, or '' if not valid.
    Returns (value, projects_for_unit_or_None, matched_how)."""
    s = _s(raw)
    up = s.upper()
    if up in ("", "N/A", "-", "NONE"):
        return "", None, "blank"
    if up in fu_vocab:
        return s, fu_vocab[up], "valid"
    # tolerate minor spacing (e.g. ' FM05-TN' -> not a Set_Up code -> invalid)
    compact = up.replace(" ", "")
    for code, projs in fu_vocab.items():
        if code.replace(" ", "") == compact:
            return code, projs, "valid"
    return "", None, "invalid"


def main():
    # locate the tracker workbook (same names ingest accepts)
    cands = [f for f in os.listdir(DATA_DIR)
             if re.search(r"(NCR_Cutover_Tracker|NCtracker|NC_Tracker).*\.xls[xm]$",
                          f, re.I) and not f.startswith("~$")]
    if not cands:
        print(f"No tracker .xlsx/.xlsm in {DATA_DIR}/"); sys.exit(1)
    path = os.path.join(DATA_DIR, sorted(cands)[0])
    print(f"tracker: {os.path.basename(path)}")

    fu_vocab = load_setup_vocab(path)
    print(f"Set_Up: {len(fu_vocab)} valid flight-unit codes, "
          f"{sum(1 for v in fu_vocab.values() if len(v)==1)} unambiguous")

    df = pd.read_excel(path, sheet_name=TRACKER_SHEET, dtype=str)
    df.columns = [_s(c) for c in df.columns]
    df = df[df.get("System").notna()].copy() if "System" in df.columns else df

    proj_col = "Project"
    fu_col = next((c for c in df.columns if c == "Flight Unit" or "Flight" in c), None)

    clean_proj, clean_fu, resolution, flags = [], [], [], []
    for _, r in df.iterrows():
        raw_p = r.get(proj_col)
        raw_f = r.get(fu_col) if fu_col else None
        p, how_p = canon_project(raw_p)
        f, projs_f, how_f = canon_flight_unit(raw_f, fu_vocab)

        notes = []
        # recover a blank/unresolved project from an UNAMBIGUOUS flight unit
        if not p and projs_f and len(projs_f) == 1:
            p = list(projs_f)[0]
            how_p = "from flight unit"
            notes.append(f"project recovered from flight unit {f}")
        # cross-check: if both known, flight unit's project must include project
        if p and projs_f and p not in projs_f:
            notes.append(f"MISMATCH: project '{p}' vs flight unit valid for {sorted(projs_f)}")
        # flags for human fix
        if how_p == "unresolved" and _s(raw_p):
            notes.append(f"unresolved project: '{_s(raw_p)}'")
        if how_f == "invalid":
            notes.append(f"invalid flight unit blanked: '{_s(raw_f)}'")
        if not p:
            notes.append("no project")

        clean_proj.append(p)
        clean_fu.append(f)
        resolution.append(f"proj:{how_p} | fu:{how_f}")
        flags.append("; ".join(notes))

    out = df.copy()
    out["Project_clean"] = clean_proj
    out["FlightUnit_clean"] = clean_fu
    out["_resolution"] = resolution
    out["_flags"] = flags

    id_col = "ID-Blackout" if "ID-Blackout" in out.columns else out.columns[1]
    show = [c for c in [id_col, "TC ID", proj_col, "Project_clean",
                        fu_col, "FlightUnit_clean", "_resolution", "_flags"]
            if c and c in out.columns]
    clean_view = out[show]
    flagged = out[out["_flags"].str.len() > 0][show]

    # summary
    n = len(out)
    changed_p = int((out[proj_col].fillna("").astype(str).str.strip()
                     != out["Project_clean"]).sum())
    recovered = int(out["_resolution"].str.contains("proj:from flight unit").sum())
    blanked_fu = int(out["_resolution"].str.contains("fu:invalid").sum())
    no_proj = int((out["Project_clean"] == "").sum())
    mism = int(out["_flags"].str.contains("MISMATCH").sum())
    summary = pd.DataFrame({
        "Metric": ["tracker rows", "Project changed by cleaning",
                   "  of which recovered from flight unit",
                   "invalid flight units blanked",
                   "rows still with NO project (need a human)",
                   "project vs flight-unit MISMATCH (need a human)",
                   "rows flagged for review (total)"],
        "Count": [n, changed_p, recovered, blanked_fu, no_proj, mism, len(flagged)],
    })

    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="summary", index=False)
        clean_view.to_excel(xw, sheet_name="clean", index=False)
        flagged.to_excel(xw, sheet_name="flags", index=False)
    _format(OUT)

    print("\n" + summary.to_string(index=False))
    print(f"\n✓ wrote {OUT}")
    print("  Review the 'flags' sheet, fix those in the tracker, re-run. "
          "Nothing has touched quality.db.")


def _format(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = load_workbook(path)
    navy = PatternFill("solid", fgColor="1E2761")
    head = Font(bold=True, color="FFFFFF")
    amber = PatternFill("solid", fgColor="FFE0B2")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.fill = navy; c.font = head; c.alignment = Alignment(vertical="center")
        for i, col in enumerate(ws.columns, 1):
            w = max((len(str(x.value)) for x in col if x.value is not None), default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 48)
        hdr = {c.value: c.column for c in ws[1]}
        fcol = hdr.get("_flags")
        if fcol:
            for row in ws.iter_rows(min_row=2, min_col=fcol, max_col=fcol):
                if _s(row[0].value):
                    row[0].fill = amber
    wb.save(path)


if __name__ == "__main__":
    main()