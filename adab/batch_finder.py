#!/usr/bin/env python3
"""Batch Finder - fill missing Charge/Batch (and as-built Revision) by lookup.

You have a NEED list (materials whose batch/revision you do NOT know - e.g. the
"RED" table) and one or more LABEL/SCAN sources that DO carry batches. For each
material in the need list this looks it up in the source(s) and pulls the
Charge (batch) + as-built Rev across.

Match key = material number (normalize-safe: type/case/space/accent invariant).
If the material is not found and BOTH sides have a description, it tries a
description match as a fallback. Each need row is flagged:
    FOUND     - exactly one batch for that material
    MULTIPLE  - several batches (candidates listed; narrow by order/serial)
    NONE      - material not in any source

Designed to become the "Find Batch" tab in ADAB. Standalone use:
    python batch_finder.py                 (opens a small picker)
    python batch_finder.py NEED SRC OUT    (command line)
"""
import os, sys, glob, unicodedata, difflib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- normalization (mirrors matchcore's rules) ----------
def norm_key(x):
    if x is None: return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit(): s = s[:-2]     # 7004369.0 -> 7004369
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.upper().replace(" ", "")

def norm_text(x):
    if x is None: return ""
    s = unicodedata.normalize("NFKD", str(x))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.upper().split())

# ---------- column detection ----------
ROLE_KEYS = {
    "material":    ["ad material", "material", "mat", "sachnummer", "sach-nr", "part", "teil", "artikel"],
    "description": ["title", "definition", "description", "bezeichnung", "kurztext", "auftragskurztext"],
    "batch":       ["charge / batch", "charge", "batch", "char", "los", "lot"],
    "rev":         ["ab rev", "rev. index", "revision", "rev"],
    "serial":      ["serialnummer", "sernr", "serial", "sn"],
    "order":       ["auftr", "auftrag", "order", "fauf", "wo"],
}
def detect(headers, role):
    hs = [("" if h is None else str(h)).strip().lower() for h in headers]
    for kw in ROLE_KEYS[role]:                 # exact-ish first
        for i, h in enumerate(hs):
            if h == kw: return i
    for kw in ROLE_KEYS[role]:                 # then contains
        for i, h in enumerate(hs):
            if kw in h and h not in ("l-char", "ch-nr"): return i
    return None

def load(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]

# ---------- core ----------
def build_index(source_paths):
    idx, didx = {}, {}
    for p in source_paths:
        hdr, rows = load(p)
        cm = {r: detect(hdr, r) for r in ROLE_KEYS}
        for r in rows:
            if cm["material"] is None: continue
            mat = norm_key(r[cm["material"]])
            if not mat: continue
            rec = {
                "batch":  r[cm["batch"]]  if cm["batch"]  is not None else None,
                "rev":    r[cm["rev"]]    if cm["rev"]    is not None else None,
                "serial": r[cm["serial"]] if cm["serial"] is not None else None,
                "order":  r[cm["order"]]  if cm["order"]  is not None else None,
                "desc":   r[cm["description"]] if cm["description"] is not None else None,
                "src":    os.path.basename(p),
            }
            idx.setdefault(mat, []).append(rec)
            d = norm_text(rec["desc"])
            if d: didx.setdefault(d, []).append(rec)
    return idx, didx

def find(need_path, source_paths):
    hdr, rows = load(need_path)
    cm = {r: detect(hdr, r) for r in ROLE_KEYS}
    idx, didx = build_index(source_paths)
    out = []
    for r in rows:
        if cm["material"] is None or r[cm["material"]] in (None, ""): continue
        mat = norm_key(r[cm["material"]])
        desc = r[cm["description"]] if cm["description"] is not None else None
        hits = list(idx.get(mat, []))
        via = "material"
        if not hits and desc:                     # description fallback
            key = norm_text(desc)
            best = difflib.get_close_matches(key, list(didx.keys()), n=1, cutoff=0.90)
            if best: hits = list(didx[best[0]]); via = "description"
        batches = sorted({str(h["batch"]).strip() for h in hits if h["batch"] not in (None, "")})
        status = "NONE" if not batches else ("FOUND" if len(batches) == 1 else "MULTIPLE")
        pick = hits[0] if len(batches) == 1 else None
        out.append({
            "row": r, "mat": r[cm["material"]], "desc": desc, "status": status,
            "batch": batches[0] if status == "FOUND" else "",
            "rev": (pick["rev"] if pick else ""), "serial": (pick["serial"] if pick else ""),
            "order": (pick["order"] if pick else ""), "src": (pick["src"] if pick else ""),
            "candidates": ", ".join(batches), "via": via if batches else "",
        })
    return hdr, cm, out

# ---------- write report ----------
def write_report(hdr, cm, results, out_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Find Batch"
    base = list(hdr)
    extra = ["Match", "Batch candidate(s)", "From order", "Matched via", "Source file"]
    head = base + extra
    GREEN = PatternFill("solid", fgColor="C6EFCE"); ORANGE = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE"); GREY = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="BFBFBF"); BORD = Border(thin, thin, thin, thin)
    for c, h in enumerate(head, 1):
        cell = ws.cell(1, c, h); cell.font = Font(name="Calibri", bold=True); cell.fill = GREY; cell.border = BORD
    fillmap = {"FOUND": GREEN, "MULTIPLE": ORANGE, "NONE": RED}
    for i, res in enumerate(results, 2):
        row = list(res["row"])
        # fill the empty as-built columns when a single batch was found
        def setcol(role, val):
            j = cm.get(role)
            if j is not None and val not in (None, ""):
                while len(row) <= j: row.append(None)
                if row[j] in (None, ""): row[j] = val
        if res["status"] == "FOUND":
            setcol("batch", res["batch"]); setcol("rev", res["rev"]); setcol("serial", res["serial"])
        vals = row + [res["status"], res["candidates"], res["order"], res["via"], res["src"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v); cell.font = Font(name="Calibri"); cell.border = BORD
            cell.fill = fillmap.get(res["status"], RED)
    for c in range(1, len(head) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(head))}{len(results)+1}"
    # summary
    s = wb.create_sheet("Summary")
    n = len(results); f = sum(r["status"]=="FOUND" for r in results)
    m = sum(r["status"]=="MULTIPLE" for r in results); z = sum(r["status"]=="NONE" for r in results)
    for i,(k,v) in enumerate([("Need rows", n), ("FOUND (1 batch)", f),
            ("MULTIPLE (pick by order/serial)", m), ("NONE (not in source)", z)], 1):
        s.cell(i,1,k).font = Font(bold=True); s.cell(i,2,v)
    s.column_dimensions["A"].width = 34
    wb.save(out_path)
    return n, f, m, z

def run(need, source, out_dir, progress=print):
    src_paths = source
    if isinstance(source, str):
        src_paths = ([source] if os.path.isfile(source)
                     else sorted(glob.glob(os.path.join(source, "*.xls*"))))
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, os.path.splitext(os.path.basename(need))[0] + "_BATCH_FOUND.xlsx")
    hdr, cm, results = find(need, src_paths)
    n, f, m, z = write_report(hdr, cm, results, out_path)
    progress(f"Need rows: {n}   FOUND: {f}   MULTIPLE: {m}   NONE: {z}")
    progress(f"Report: {out_path}")
    return out_path, (n, f, m, z)

# ---------- optional tiny GUI (ADAB-style) ----------
def gui():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.title("Batch Finder"); root.geometry("820x520")
    need = tk.StringVar(); src = tk.StringVar(); out = tk.StringVar()
    def row(r, lab, var, folder=False):
        tk.Label(root, text=lab, font=("Segoe UI", 9, "bold")).grid(row=r, column=0, columnspan=3, sticky="w", padx=12, pady=(8,0))
        tk.Entry(root, textvariable=var).grid(row=r+1, column=0, sticky="ew", padx=(12,6))
        if not folder:
            tk.Button(root, text="File...", command=lambda: var.set(filedialog.askopenfilename() or var.get())).grid(row=r+1, column=1)
        tk.Button(root, text="Folder...", command=lambda: var.set(filedialog.askdirectory() or var.get())).grid(row=r+1, column=2, padx=(0,12))
    row(0, "Need-batch list (e.g. RED):", need)
    row(2, "Label / scan source (file or folder):", src)
    row(4, "Output folder:", out, folder=True)
    log = tk.Text(root, height=14, bg="#111", fg="#ddd", font=("Consolas", 9))
    log.grid(row=8, column=0, columnspan=3, sticky="nsew", padx=12, pady=8)
    def go():
        log.delete("1.0","end")
        try: run(need.get().strip(), src.get().strip(), out.get().strip(), lambda m:(log.insert("end",str(m)+"\n"), log.see("end")))
        except Exception as e: log.insert("end", "ERROR: "+str(e)+"\n")
    tk.Button(root, text="Run", bg="#2d7d46", fg="white", font=("Segoe UI",12,"bold"), height=2, command=go).grid(row=7, column=0, columnspan=3, sticky="ew", padx=12, pady=8)
    root.grid_columnconfigure(0, weight=1); root.grid_rowconfigure(8, weight=1)
    root.mainloop()

if __name__ == "__main__":
    if len(sys.argv) >= 4:
        run(sys.argv[1], sys.argv[2], sys.argv[3])
    else:
        try: gui()
        except Exception: print("Usage: python batch_finder.py NEED SOURCE OUTDIR")