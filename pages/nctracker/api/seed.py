"""Load the Excel tracker into the database. Safe to re-run: it refuses to
overwrite a database that already holds NCs unless --force is given."""
from __future__ import annotations

import os
import sys
import pathlib

import openpyxl
import psycopg

import mapping as M

DSN = os.environ["DATABASE_URL"]
XLSX = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                    else "/app/seed/NCR_Cutover_Tracker.xlsx")
FORCE = "--force" in sys.argv


def read_setup(wb) -> dict:
    ws = wb["Set_Up"]
    rows = [[M.norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    out: dict = {}
    for ci, head in enumerate(rows[0]):
        if not head:
            continue
        seen = []
        for r in rows[1:]:
            v = r[ci] if ci < len(r) else ""
            if v and v not in seen:
                seen.append(v)          # deduplicated on load
        out[head] = seen
    # a project always has a real flight unit
    for key in ("Ariane", "Vega", "MHI H3", "Relativity", "SAS", "Vulcan", "Flexline"):
        out[key] = [v for v in out.get(key, []) if v.lower() not in ("n/a", "na")]
    out["Classification"] = ["Minor", "Major"]      # levels dropped on purpose
    out.pop("Supplier name", None)                  # supplier is free text
    return out


def read_tracker(wb) -> list:
    ws = wb["NC_Tracker_Black_Out"]
    raw = list(ws.iter_rows(values_only=True))
    cmap = M.map_headers([M.norm(c) for c in raw[0]])
    return [M.clean_row({db: r[i] for i, db in cmap.items() if i < len(r)})
            for r in raw[1:] if any(M.norm(c) for c in r)]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    setup, rows = read_setup(wb), read_tracker(wb)

    with psycopg.connect(DSN) as c, c.cursor() as cur:
        cur.execute(pathlib.Path(__file__).with_name("schema.sql").read_text())
        cur.execute("SELECT count(*) FROM nc")
        if cur.fetchone()[0] and not FORCE:
            print("database already holds NCs — nothing done. Use --force to reload.")
            return
        if FORCE:
            cur.execute("TRUNCATE nc, setup_value, audit RESTART IDENTITY")

        for name, values in setup.items():
            for i, v in enumerate(values):
                cur.execute("""INSERT INTO setup_value (list_name, value, sort_order)
                               VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""", (name, v, i))

        cols = M.DB_COLS + ["match_key"]
        holes = ", ".join(["%s"] * len(cols))
        for r in rows:
            cur.execute(f"INSERT INTO nc ({', '.join(cols)}) VALUES ({holes})",
                        [r.get(col, "") for col in M.DB_COLS] + [M.match_key(r)])

        # the file itself, so "old Excel version" can be downloaded later
        cur.execute("INSERT INTO upload (kind, filename, content) VALUES (%s,%s,%s)",
                    ("old_excel", XLSX.name, XLSX.read_bytes()))
        c.commit()

    print(f"seeded {len(rows)} NCs and "
          f"{sum(len(v) for v in setup.values())} set-up values from {XLSX.name}")


if __name__ == "__main__":
    main()