"""NC Tracker API.

Serves the single-page front end and the JSON it needs. Runs in the nct-api
container; the database is only reachable from inside the nct-net network.
"""
from __future__ import annotations

import io
import os
import datetime as dt

import openpyxl
import psycopg
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from psycopg.rows import dict_row
from pydantic import BaseModel

import mapping as M

DSN = os.environ["DATABASE_URL"]
WEB = os.environ.get("WEB_DIR", "/app/web")
UPLOAD_KINDS = {"nc_tracker": "NC Tracker workbook",
                "tc_report": "TC report",
                "nc_sap_overview": "NC SAP overview",
                "old_excel": "NC Tracker — old Excel version"}
# a TC report is the EZ1 landscape, a SAP overview is SAP. Used only when the
# file itself carries no System column.
#
# nc_tracker is deliberately absent: the tracker workbook carries its own
# System column on every row, so stamping a fallback would rewrite Blackout
# and SAP rows as EZ1. A blank System stays blank and shows up in the
# "no system" count, which is the honest outcome.
DEFAULT_SYSTEM = {"tc_report": "EZ1", "nc_sap_overview": "SAP"}

app = FastAPI(title="NC Tracker")


def db():
    return psycopg.connect(DSN, row_factory=dict_row)


def load_setup(cur) -> dict:
    cur.execute("SELECT list_name, value FROM setup_value ORDER BY list_name, sort_order, id")
    out: dict = {}
    for r in cur.fetchall():
        out.setdefault(r["list_name"], []).append(r["value"])
    return out


def load_rows(cur) -> list:
    cur.execute(f"SELECT id, {', '.join(M.DB_COLS)} FROM nc ORDER BY id")
    return cur.fetchall()


def load_capa(cur) -> list:
    cur.execute(f"SELECT id, {', '.join(M.CAPA_DB_COLS)} FROM capa ORDER BY id")
    return cur.fetchall()


# --------------------------------------------------------------------------
# read
# --------------------------------------------------------------------------
@app.get("/api/bootstrap")
def bootstrap():
    with db() as c, c.cursor() as cur:
        setup, rows, capa = load_setup(cur), load_rows(cur), load_capa(cur)
        cur.execute("""SELECT DISTINCT ON (kind) kind, filename, uploaded_at
                       FROM upload ORDER BY kind, uploaded_at DESC""")
        uploads = {r["kind"]: {"filename": r["filename"],
                               "uploaded_at": r["uploaded_at"].isoformat()}
                   for r in cur.fetchall()}
    return {"headers": M.HEADERS, "columns": M.COLUMNS, "rows": rows,
            "capa_headers": M.CAPA_HEADERS, "capa_columns": M.CAPA_COLUMNS,
            "capa_rows": capa, "capa_list_of": M.CAPA_LIST_OF,
            "capa_l2_parent": M.CAPA_L2_PARENT,
            "setup": setup, "uploads": uploads}


# --------------------------------------------------------------------------
# edit one cell
# --------------------------------------------------------------------------
class Edit(BaseModel):
    field: str          # database column
    value: str = ""


@app.patch("/api/nc/{nc_id}")
def edit(nc_id: int, body: Edit):
    if body.field not in M.DB_COLS:
        raise HTTPException(400, f"unknown field '{body.field}'")
    value = M.clean_value(body.field, body.value)
    with db() as c, c.cursor() as cur:
        cur.execute(f"SELECT id, {', '.join(M.DB_COLS)} FROM nc WHERE id=%s", (nc_id,))
        before = cur.fetchone()
        if before is None:
            raise HTTPException(404, "no such NC")
        row = {"old": before[body.field]}
        cur.execute(
            f"UPDATE nc SET {body.field}=%s, updated_at=now() WHERE id=%s "
            f"RETURNING id, {', '.join(M.DB_COLS)}", (value, nc_id))
        updated = cur.fetchone()
        cur.execute("""INSERT INTO audit (nc_id, field, old_value, new_value, source)
                       VALUES (%s,%s,%s,%s,'edit')""",
                    (nc_id, body.field, row["old"], value))

        # A real TC ID typed onto a Burndown row means the NC has migrated:
        # it moves to EZ1 straight away, and the old System is kept in the audit.
        if body.field == "tc_id":
            move = M.migration_update(before, value)
            for field, new_val in move.items():
                cur.execute(f"UPDATE nc SET {field}=%s WHERE id=%s", (new_val, nc_id))
                cur.execute("""INSERT INTO audit (nc_id, field, old_value, new_value, source)
                               VALUES (%s,%s,%s,%s,'auto-migrate')""",
                            (nc_id, field, M.norm(before.get(field)), new_val))
                updated[field] = new_val
        key = M.match_key(updated)
        cur.execute("UPDATE nc SET match_key=%s WHERE id=%s", (key, nc_id))
        c.commit()
    return updated


@app.patch("/api/capa/{capa_row_id}")
def capa_edit(capa_row_id: int, body: Edit):
    if body.field not in M.CAPA_DB_COLS:
        raise HTTPException(400, f"unknown field '{body.field}'")
    if M.CAPA_KIND.get(body.field) == "calc":
        raise HTTPException(400, f"'{body.field}' is calculated, not typed")
    value = M.capa_clean(body.field, body.value)
    with db() as c, c.cursor() as cur:
        cur.execute(f"SELECT {body.field} AS old FROM capa WHERE id=%s", (capa_row_id,))
        row = cur.fetchone()
        if row is None:
            raise HTTPException(404, "no such CAPA")
        cur.execute(f"UPDATE capa SET {body.field}=%s, updated_at=now() WHERE id=%s "
                    f"RETURNING id, {', '.join(M.CAPA_DB_COLS)}", (value, capa_row_id))
        updated = cur.fetchone()
        # status and days open follow the dates; they are never typed
        updated["status"] = M.capa_status(updated)
        updated["days_open"] = M.capa_days_open(updated)
        cur.execute("UPDATE capa SET status=%s, days_open=%s WHERE id=%s",
                    (updated["status"], updated["days_open"], capa_row_id))
        cur.execute("""INSERT INTO audit (nc_id, field, old_value, new_value, source)
                       VALUES (%s,%s,%s,%s,'capa-edit')""",
                    (capa_row_id, body.field, row["old"], value))
        c.commit()
    return updated


@app.post("/api/capa")
def capa_add():
    """A new, empty CAPA line. Everything shows yellow until it is filled in."""
    today = dt.date.today().isoformat()
    with db() as c, c.cursor() as cur:
        cur.execute("INSERT INTO capa (open_date, status) VALUES (%s,'Open') "
                    f"RETURNING id, {', '.join(M.CAPA_DB_COLS)}", (today,))
        row = cur.fetchone()
        c.commit()
    return row


# --------------------------------------------------------------------------
# set-up lists
# --------------------------------------------------------------------------
class SetupAdd(BaseModel):
    list_name: str
    value: str


@app.post("/api/setup")
def setup_add(body: SetupAdd):
    value = M.norm(body.value)
    if not value:
        raise HTTPException(400, "empty value")
    with db() as c, c.cursor() as cur:
        cur.execute("""INSERT INTO setup_value (list_name, value, sort_order)
                       VALUES (%s,%s,999) ON CONFLICT DO NOTHING""",
                    (body.list_name, value))
        c.commit()
        return load_setup(cur)


@app.delete("/api/setup")
def setup_remove(list_name: str, value: str):
    with db() as c, c.cursor() as cur:
        cur.execute("DELETE FROM setup_value WHERE list_name=%s AND value=%s",
                    (list_name, value))
        c.commit()
        return load_setup(cur)


# --------------------------------------------------------------------------
# import — preview first, commit second. Nothing is written by the preview.
# --------------------------------------------------------------------------
def read_upload(content: bytes, kind: str = "", projects: list | None = None) -> tuple:
    """Read every sheet that looks like NC data, not just the best one.

    The tracker's own export has Burndown AND EZ1 as separate sheets. Reading
    only the highest-scoring sheet silently dropped the other one, so a 241-row
    export imported as 185. Any sheet within 4 columns of the best is read.
    """
    # Two ways to read the same file, and both are needed.
    #
    #   read_only=True   streams the sheet. Fast, low memory, and the only way
    #                    a hand-maintained Excel with styling on thousands of
    #                    empty rows opens in reasonable time.
    #   read_only=False  materialises every cell. Slow, but it is the only mode
    #                    that recovers SAP S/4 exports, which declare their
    #                    dimensions wrongly and stream as a single cell.
    #
    # So: stream first, and fall back only when streaming clearly failed.
    def sheets_from(workbook) -> list:
        found = []
        for nm in workbook.sheetnames:
            rws = []
            for i, r in enumerate(workbook[nm].iter_rows(values_only=True)):
                if i >= MAX_ROWS:
                    break
                rws.append(r)
            # drop the trailing block of formatted-but-empty rows
            while rws and not any(M.norm(c) for c in rws[-1]):
                rws.pop()
            if rws:
                found.append((nm, rws))
        return found

    MAX_ROWS = 20000
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        pages = sheets_from(wb)
    finally:
        wb.close()

    # A single row, or nothing at all, means the file lied about its size.
    if sum(len(r) for _, r in pages) < 2:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        pages = sheets_from(wb)

    seen_headers: list = []
    sheets = []                       # (score, sheet name, rows, header index, column map)

    for name, raw in pages:
        if not raw:
            continue
        best_here = (0, 0, {})
        for i, row in enumerate(raw[:15]):
            cmap = M.map_headers([M.norm(c) for c in row])
            if len(cmap) > best_here[0]:
                best_here = (len(cmap), i, cmap)
            if i == 0:
                seen_headers += [M.norm(c) for c in row if M.norm(c)]
        if best_here[0]:
            sheets.append((best_here[0], name, raw, best_here[1], best_here[2]))

    if not sheets:
        raise HTTPException(400, "the file has no readable sheet")

    top = max(s[0] for s in sheets)
    if top < 3:
        raise HTTPException(
            400, "no tracker columns recognised in this file. Columns found: "
                 + (", ".join(seen_headers[:25]) or "none"))

    # A sheet counts as NC data when it maps nearly as many columns as the best
    # one. The CAPA board (6) and Set-up (1) fall far below and stay out.
    keep = [s for s in sheets if s[0] >= max(3, top - 4)]

    rows, matched = [], set()
    for score, name, raw, head_i, cmap in keep:
        matched |= set(cmap.values())
        rows += [M.clean_row({db: r[i] for i, db in cmap.items() if i < len(r)})
                 for r in raw[head_i + 1:] if any(M.norm(c) for c in r)]

    fallback = DEFAULT_SYSTEM.get(kind, "")
    for row in rows:
        if fallback and not row.get("system"):
            row["system"] = fallback
        # a Teamcenter contract code becomes the project name the tracker uses
        if projects and row.get("project"):
            row["project"] = M.map_project(row["project"], projects)
    return rows, sorted(matched)


@app.post("/api/import/{kind}/preview")
async def import_preview(kind: str, file: UploadFile = File(...)):
    if kind not in UPLOAD_KINDS:
        raise HTTPException(400, f"unknown import '{kind}'")
    with db() as c, c.cursor() as cur:
        setup = load_setup(cur)
    incoming, matched = read_upload(await file.read(), kind, setup.get("Project", []))
    with db() as c, c.cursor() as cur:
        result = M.diff(load_rows(cur), incoming, setup)
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT count(*) AS n FROM nc")
        existing_total = cur.fetchone()["n"]
    result.update({"filename": file.filename, "rows_in_file": len(incoming),
                   "columns_recognised": matched, "kind": kind,
                   "label": UPLOAD_KINDS[kind],
                   "tracker_before": existing_total,
                   "tracker_after": existing_total + result["new"]})
    return result


@app.post("/api/import/{kind}/commit")
async def import_commit(kind: str, file: UploadFile = File(...)):
    if kind not in UPLOAD_KINDS:
        raise HTTPException(400, f"unknown import '{kind}'")
    content = await file.read()
    with db() as c, c.cursor() as cur:
        projects = load_setup(cur).get("Project", [])
    incoming, _ = read_upload(content, kind, projects)
    added = updated = 0
    with db() as c, c.cursor() as cur:
        cur.execute("INSERT INTO upload (kind, filename, content) VALUES (%s,%s,%s)",
                    (kind, file.filename, content))
        existing = {r["match_key"] or M.match_key(r): r for r in load_rows_full(cur)}
        for row in incoming:
            key = M.match_key(row)
            old = existing.get(key)
            if old is None:
                cur.execute(
                    f"INSERT INTO nc ({', '.join(M.DB_COLS)}, match_key) "
                    f"VALUES ({', '.join(['%s'] * len(M.DB_COLS))}, %s)",
                    [row.get(c, "") for c in M.DB_COLS] + [key])
                added += 1
                continue
            # Never overwrite a value a person has typed. Only fill blanks.
            # Rule: incoming has content AND the tracker cell is empty.
            delta = {c: row[c] for c in M.DB_COLS
                     if row.get(c) and not M.norm(old.get(c))}
            if not delta:
                continue
            sets = ", ".join(f"{c}=%s" for c in delta)
            cur.execute(f"UPDATE nc SET {sets}, updated_at=now() WHERE id=%s",
                        list(delta.values()) + [old["id"]])
            for field, new in delta.items():
                cur.execute("""INSERT INTO audit (nc_id, field, old_value, new_value, source)
                               VALUES (%s,%s,%s,%s,%s)""",
                            (old["id"], field, M.norm(old.get(field)), new, f"import:{kind}"))
            updated += 1
        c.commit()
    return {"added": added, "updated": updated, "stored_as": file.filename}


def load_rows_full(cur) -> list:
    cur.execute(f"SELECT id, match_key, {', '.join(M.DB_COLS)} FROM nc")
    return cur.fetchall()


# --------------------------------------------------------------------------
# downloads
# --------------------------------------------------------------------------
@app.get("/api/uploads/status")
def uploads_status():
    """Latest upload time and filename for every kind — for the small line
    under the page title. Never fails: unseen kinds come back as null."""
    with db() as c, c.cursor() as cur:
        cur.execute("""SELECT DISTINCT ON (kind) kind, filename, uploaded_at
                       FROM upload
                       ORDER BY kind, uploaded_at DESC""")
        rows = cur.fetchall()
    out = {k: None for k in UPLOAD_KINDS}
    for r in rows:
        out[r["kind"]] = {"filename": r["filename"],
                          "uploaded_at": r["uploaded_at"].isoformat(),
                          "label": UPLOAD_KINDS.get(r["kind"], r["kind"])}
    return out


@app.get("/api/upload/{kind}/latest")
def upload_latest(kind: str):
    if kind not in UPLOAD_KINDS:
        raise HTTPException(400, f"unknown file '{kind}'")
    with db() as c, c.cursor() as cur:
        cur.execute("""SELECT filename, content FROM upload WHERE kind=%s
                       ORDER BY uploaded_at DESC LIMIT 1""", (kind,))
        row = cur.fetchone()
    if row is None:
        raise HTTPException(404, f"no {UPLOAD_KINDS[kind]} has been uploaded yet")
    return Response(
        row["content"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"'})


# The page's tabs, and which System values each one holds. Burndown is SAP and
# Blackout together, in one sheet — they are never shown separately.
TABS = {"Burndown": ["SAP", "Blackout"], "EZ1": ["EZ1"]}


@app.get("/api/export")
def export(scope: str = "all", tab: str = ""):
    """scope=all -> every tab in one workbook. scope=tab -> that tab only."""
    with db() as c, c.cursor() as cur:
        rows, setup = load_rows(cur), load_setup(cur)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wanted = list(TABS) if scope == "all" else [tab if tab in TABS else "Burndown"]

    for tab_name in wanted:
        systems = TABS[tab_name]
        ws = wb.create_sheet(tab_name[:31])
        ws.append(M.HEADERS)
        for r in rows:
            if M.norm(r.get("system")) not in systems:
                continue
            ws.append([M.dmy(r.get(db_col)) if kind == "date" else M.norm(r.get(db_col))
                       for _, db_col, kind in M.COLUMNS])

    if scope == "all":
        with db() as c, c.cursor() as cur:
            capa = load_capa(cur)
        ws = wb.create_sheet("CAPA board")
        ws.append(M.CAPA_HEADERS)
        for r in capa:
            if M.norm(r.get("status")) == "Closed":
                continue                      # the board is open work only
            ws.append([M.dmy(r.get(db_col)) if kind == "date" else M.norm(r.get(db_col))
                       for _, db_col, kind in M.CAPA_COLUMNS])

        ws = wb.create_sheet("Set-up")
        ws.append(["List", "Value"])
        for name, values in setup.items():
            for v in values:
                ws.append([name, v])

        ws = wb.create_sheet("NC Report")
        ws.append(["Issue owner", "Closed", "Open", "Not ticked", "Total"])
        owners = sorted({M.norm(r.get("owner")) or "(not set)" for r in rows})
        for o in owners:
            sub = [r for r in rows if (M.norm(r.get("owner")) or "(not set)") == o]
            ws.append([o,
                       sum(1 for r in sub if M.norm(r.get("status")) == "Closed"),
                       sum(1 for r in sub if M.norm(r.get("status")) == "Open"),
                       sum(1 for r in sub if M.norm(r.get("status")) not in ("Open", "Closed")),
                       len(sub)])

    buf = io.BytesIO()
    wb.save(buf)
    stamp = dt.date.today().strftime("%Y-%m-%d")
    tag = "all_tabs" if scope == "all" else wanted[0]
    name = f"NC_Tracker_{tag}_{stamp}.xlsx"
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'})


# --------------------------------------------------------------------------
# front end
# --------------------------------------------------------------------------
@app.post("/api/admin/reset-nc")
def reset_nc(confirm: str = ""):
    """Delete every NC row, its audit trail and the stored uploads.

    CAPA rows and the Set-up lists are NOT touched — they are separate work.
    Requires ?confirm=DELETE-ALL-NC so it can never fire by accident."""
    if confirm != "DELETE-ALL-NC":
        raise HTTPException(400, "add ?confirm=DELETE-ALL-NC to run this")
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT count(*) AS n FROM nc")
        before = cur.fetchone()["n"]
        cur.execute("DELETE FROM audit WHERE nc_id IN (SELECT id FROM nc)")
        cur.execute("DELETE FROM upload")
        cur.execute("DELETE FROM nc")
        cur.execute("SELECT count(*) AS n FROM capa")
        capa_left = cur.fetchone()["n"]
        cur.execute("SELECT count(*) AS n FROM setup_value")
        setup_left = cur.fetchone()["n"]
        c.commit()
    return {"deleted_ncs": before, "capa_rows_untouched": capa_left,
            "setup_values_untouched": setup_left}


@app.get("/health")
def health():
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT count(*) AS n FROM nc")
        return {"ok": True, "ncs": cur.fetchone()["n"]}


@app.get("/")
def index():
    return FileResponse(os.path.join(WEB, "index.html"))


@app.exception_handler(HTTPException)
def http_error(_, exc: HTTPException):
    return JSONResponse({"error": exc.detail}, status_code=exc.status_code)