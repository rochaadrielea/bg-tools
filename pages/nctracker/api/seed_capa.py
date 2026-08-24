"""Load the CAPA set-up lists from USE_THIS_CAPA_2.csv into the database.

The CSV is a vocabulary sheet, not a list of CAPAs: one column per dropdown.
Every list is stored in the same `setup_value` table as the NC lists, with the
name prefixed `capa:` so the two set-ups stay apart.

    python seed_capa.py /app/seed/USE_THIS_CAPA_2.csv
"""
from __future__ import annotations

import csv
import os
import pathlib
import sys

import psycopg

import mapping as M

DSN = os.environ["DATABASE_URL"]
CSV_PATH = pathlib.Path(sys.argv[1] if len(sys.argv) > 1
                        else "/app/seed/USE_THIS_CAPA_2.csv")

# column index in the CSV -> list name shown in Set-up.
# Indexes come from reading the file; blank spacer columns are skipped.
COLUMNS = {
    0:  "Priority",
    2:  "Status",
    4:  "NC Type",
    6:  "Origin",
    8:  "Request type",
    10: "Type",
    11: "Risk level",
    12: "Target (days)",
    15: "Risk level (result)",
    20: "L1 Origin Area",
    23: "L2 Bonding",
    24: "L2 Integration",
    25: "L2 Machining",
    26: "L2 Completion",
    27: "L2 Supplier",
    28: "L2 Test",
    29: "L2 Warehouse",
    30: "L2 Customer",
    31: "L2 Incoming Inspection",
    37: "L1 Root cause",
    39: "L2 Documentation",
    40: "L2 Material",
    41: "L2 People",
    42: "L2 Tool",
    51: "Responsible area",
    53: "Classification",
    55: "MAIT flow impacted in next 5 days",
    57: "DRB planned in 1 month",
    59: "Implementation Verification",
    64: "Responsible",
    67: "Project Manager",
    72: "Affected Project",
    76: "Department Assigned",
}

# values that mean "nothing here" and never belong on a list
JUNK = {"#n/a", "na", "n/a", ""}


def read_lists(path: pathlib.Path) -> dict:
    with path.open(encoding="utf-8-sig", errors="replace", newline="") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        raise SystemExit(f"{path.name} is empty")
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]

    out: dict = {}
    for ci, name in COLUMNS.items():
        if ci >= width:
            continue
        seen: list = []
        for r in rows[1:]:
            v = M.norm(r[ci])
            if v.lower() in JUNK or v in seen:
                continue
            seen.append(v)                       # deduplicated on load
        if seen:
            out[name] = seen
    return out


BOARD_SHEET = "Capa Board "        # the trailing space is in the workbook


def read_board(path: pathlib.Path) -> list:
    """Every row of the Capa Board, cleaned. Nothing is dropped here — the
    page decides which statuses to show."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == "capa board"), None)
    if sheet is None:
        raise SystemExit(f"no 'Capa Board' sheet in {path.name}; found {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    head = [M.norm(h) for h in rows[0]]
    # position-based: the board's column order is the contract
    out = []
    for r in rows[1:]:
        if not any(M.norm(c) for c in r):
            continue
        raw = {db: (r[i] if i < len(r) else "")
               for i, (_, db, _) in enumerate(M.CAPA_COLUMNS)}
        row = M.capa_clean_row(raw)
        # an Excel table runs on past its data; a row with nothing that
        # identifies a CAPA is padding, not a record
        if not any(row.get(c) for c in ("capa_id", "problem", "nc_number",
                                        "requestor", "responsible")):
            continue
        out.append(row)
    return out


def lists_from_board(board: list) -> dict:
    """Lists the CSV does not carry but the board uses."""
    def distinct(col):
        seen = []
        for r in board:
            v = M.norm(r.get(col))
            if v and v.lower() not in JUNK and v not in seen:
                seen.append(v)
        return sorted(seen)
    return {
        "Supplier": distinct("supplier"),
        "Department Responsible": distinct("dept_responsible"),
        "ID Number": distinct("id_number"),
        "Yes/No": ["Yes", "No", "N/A"],
    }


def main() -> None:
    lists = read_lists(CSV_PATH)

    board = []
    xlsm = next((p for p in CSV_PATH.parent.glob("*.xlsm")), None)
    if xlsm:
        board = read_board(xlsm)
        lists.update(lists_from_board(board))

    with psycopg.connect(DSN) as c, c.cursor() as cur:
        cur.execute(pathlib.Path(__file__).with_name("schema.sql").read_text())
        cur.execute("DELETE FROM setup_value WHERE list_name LIKE 'capa:%'")
        for name, values in lists.items():
            for i, v in enumerate(values):
                cur.execute("""INSERT INTO setup_value (list_name, value, sort_order)
                               VALUES (%s,%s,%s) ON CONFLICT DO NOTHING""",
                            (f"capa:{name}", v, i))
        if board:
            cur.execute("TRUNCATE capa RESTART IDENTITY")
            cols = M.CAPA_DB_COLS
            holes = ", ".join(["%s"] * len(cols))
            for r in board:
                cur.execute(f"INSERT INTO capa ({', '.join(cols)}) VALUES ({holes})",
                            [r.get(col, "") for col in cols])
        c.commit()
    if board:
        from collections import Counter
        print(f"loaded {len(board)} CAPA rows from {xlsm.name}")
        for k, v in Counter(r["status"] for r in board).most_common():
            print(f"  status {k or '(not set)':12} {v}")
    total = sum(len(v) for v in lists.values())
    print(f"loaded {len(lists)} CAPA lists, {total} values from {CSV_PATH.name}")
    for name, values in lists.items():
        print(f"  {name:36} {len(values):3}")


if __name__ == "__main__":
    main()