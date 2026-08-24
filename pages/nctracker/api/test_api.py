"""End-to-end test. Needs DATABASE_URL pointing at a seeded database.

    DATABASE_URL=postgresql://nct@127.0.0.1:55432/nctracker python3 test_api.py
"""
import io
import json
import openpyxl
from fastapi.testclient import TestClient

import main

c = TestClient(main.app)


def show(label, r):
    body = r.json() if r.headers.get("content-type", "").startswith("application/json") else r.content
    print(f"{label:34} {r.status_code}  "
          + (json.dumps(body)[:150] if isinstance(body, (dict, list)) else f"{len(body)} bytes"))
    return body


b = show("bootstrap", c.get("/api/bootstrap"))
print("   rows", len(b["rows"]), "| setup lists", len(b["setup"]),
      "| Ariane has n/a:", "n/a" in b["setup"]["Ariane"],
      "| Classification", b["setup"]["Classification"])

show("edit status -> closed", c.patch("/api/nc/2", json={"field": "status", "value": "closed"}))
show("edit unknown field (must 400)", c.patch("/api/nc/2", json={"field": "nope", "value": "x"}))
show("edit missing NC (must 404)", c.patch("/api/nc/99999", json={"field": "status", "value": "Open"}))
show("add set-up value", c.post("/api/setup", json={"list_name": "Ariane", "value": "C6001L T-Half"}))

# a TC report with completely different headers
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "TC export"
ws.append(["Issue Number", "Type", "Programme", "Unit", "Detection Area",
           "Short Text", "Owner", "Creation Date", "Severity", "System status"])
ws.append(["IR-009001", "Internal", "Ariane", "A6017", "B: Bonding",
           "new bonding finding", "Vitaly Meshin", "2026-08-01", "Minor", "Open"])
ws.append(["IR-009002", "Internal", "SAS", "Batch 22", "T: Testing",
           "new test finding", "Elisa Martin", "2026-08-02", "Major", "Open"])
ws.append(["IR-001382", "Internal", "Ariane", "A6017", "B: Bonding",
           "title changed by the TC report", "Vitaly Meshin", "2026-07-16", "Minor", "Closed"])
buf = io.BytesIO()
wb.save(buf)
tc = ("tc_report.xlsx", buf.getvalue(),
      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

p = show("import preview", c.post("/api/import/tc_report/preview", files={"file": tc}))
print("   new", p["new"], p["new_by_system"], "| on EZ1", p["new_on_ez1"],
      "| updated", p["updated"], "| unchanged", p["unchanged"], "| flagged", p["flagged"])
print("   recognised", p["columns_recognised"])
if p["changes"]:
    print("   change:", p["changes"][0]["fields"])

before = c.get("/health").json()["ncs"]
show("import commit", c.post("/api/import/tc_report/commit", files={"file": tc}))
after = c.get("/health").json()["ncs"]
print("   NCs", before, "->", after)

bad = io.BytesIO()
wbb = openpyxl.Workbook()
wbb.active.append(["alpha", "beta", "gamma"])
wbb.save(bad)
show("import junk (must 400)", c.post(
    "/api/import/tc_report/preview",
    files={"file": ("junk.xlsx", bad.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}))

x = show("export all tabs", c.get("/api/export?scope=all"))
wb2 = openpyxl.load_workbook(io.BytesIO(x))
print("   sheets", wb2.sheetnames, "| SAP rows", wb2["SAP"].max_row,
      "| date cell", wb2["SAP"].cell(2, 14).value)
x = show("export one tab", c.get("/api/export?scope=tab&system=EZ1"))
print("   sheets", openpyxl.load_workbook(io.BytesIO(x)).sheetnames)

show("download old excel", c.get("/api/upload/old_excel/latest"))
show("download tc report", c.get("/api/upload/tc_report/latest"))
show("download never uploaded (must 404)", c.get("/api/upload/nc_sap_overview/latest"))
print("\nall checks ran")