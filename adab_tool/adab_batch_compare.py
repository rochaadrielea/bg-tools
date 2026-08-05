#!/usr/bin/env python3
"""
ADAB Batch Comparator  -  As-Designed (MBOM) vs As-Built, serial by serial.

Replicates the ADAB_Standard_Report.xlsm macro (Merge_AsDesigned_AsBuild):
  - joins each As-Built unit against the single As-Designed baseline on TCUID child token
  - builds one sheet: MBOM column block | As-Build column block | Status col A
  - classifies each row and colours it exactly like the macro:
        Assembled                        -> light green
        Missing / Missing Part           -> light red
        Deviated* (added / rev changed)  -> light orange

  ONE REPORT FILE PER UNIT. Each As-Built produces its own workbook,
  named <prefix>_<SERIAL>.xlsx, written next to where the script is run.

Usage:
    python adab_batch_compare.py --design DESIGN_FILE_OR_FOLDER \
                                 --built BUILT_FOLDER \
                                 --prefix ADAB_Report

The design folder must contain exactly one file (the F- baseline).
The built folder contains one .xlsm per serial (CH0960, CH0984, ...).
"""

import argparse
import glob
import os
import re
import warnings
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Silence openpyxl's harmless ">31 characters" sheet-title warning. Long
# assembly names (e.g. "Kinematic Instr. Installation T-Half") trip it on load;
# Excel just truncates the tab name. Nothing is lost, so we hide the noise.
warnings.filterwarnings(
    "ignore", message=".*more than 31 characters.*", category=UserWarning)

# --- matchcore: the reusable matching engine (material key + description hash +
#     TF-IDF vector). Optional at import time so the tool still runs if the
#     package is missing, but it lives in this folder and is normally present. ---
try:
    from matchcore import (norm_key as mc_norm_key, desc_hash as mc_desc_hash,
                           record_hash as mc_record_hash,
                           build_vectorizer as mc_build_vec, agree as mc_agree)
    _HAS_MATCHCORE = True
except Exception:
    _HAS_MATCHCORE = False

# --- SELF-CONTAINED description-witness fallback ----------------------------
# The description engine (normalise -> TF-IDF vectorise -> cosine compare, plus
# an exact hash) is the second witness Adriele asked to ALWAYS see running. If
# the matchcore package can't be imported on a given machine, we run a byte-for-
# byte-equivalent copy here so the vectorisation still runs and the "Match
# Engine" columns always appear. Same public names as matchcore, so the rest of
# the engine doesn't care which one is active.
if not _HAS_MATCHCORE:
    import math as _math, hashlib as _hl, unicodedata as _ud
    from collections import Counter as _Counter

    def mc_norm_key(v):
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return re.sub(r"\s+", " ", str(v).strip().upper())

    def _norm_text(v):
        if v is None:
            return ""
        s = _ud.normalize("NFKC", str(v))
        s = "".join(c for c in s if not _ud.combining(c)).casefold()
        s = re.sub(r"[^a-z0-9]+", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    def mc_desc_hash(d):
        return _hl.sha256(("text:" + _norm_text(d)).encode("utf-8")).hexdigest()

    def mc_record_hash(material, description, batch=None):
        keys = [material] if batch is None else [material, batch]
        payload = "␟".join(mc_norm_key(k) for k in keys) + \
                  "␟" + "text:" + _norm_text(description)
        return _hl.sha256(payload.encode("utf-8")).hexdigest()

    class _Vec:
        def __init__(self, corpus):
            docs = [_norm_text(c).split() for c in corpus]
            df = _Counter()
            for d in docs:
                for t in set(d):
                    df[t] += 1
            n = max(1, len(docs))
            self.idf = {t: _math.log((1 + n) / (1 + c)) + 1.0 for t, c in df.items()}
            self._d = _math.log((1 + n) / 1) + 1.0

        def vec(self, text):
            tf = _Counter(_norm_text(text).split())
            return {t: c * self.idf.get(t, self._d) for t, c in tf.items()}

    def _cosine(a, b):
        if not a or not b:
            return 0.0
        dot = sum(w * b[t] for t, w in a.items() if t in b)
        na = _math.sqrt(sum(w * w for w in a.values()))
        nb = _math.sqrt(sum(w * w for w in b.values()))
        return (dot / (na * nb)) if (na and nb) else 0.0

    def mc_build_vec(*desc_iterables):
        corpus = []
        for it in desc_iterables:
            corpus.extend(it)
        return _Vec(corpus)

    class _Agreement:
        def __init__(self, material_equal, desc_hash_equal, desc_similarity, verdict):
            self.material_equal = material_equal
            self.desc_hash_equal = desc_hash_equal
            self.desc_similarity = desc_similarity
            self.verdict = verdict

    def mc_agree(mat_a, desc_a, mat_b, desc_b, vec, sim_threshold=0.60):
        ka, kb = mc_norm_key(mat_a), mc_norm_key(mat_b)
        material_equal = bool(ka) and ka == kb
        ta, tb = _norm_text(desc_a), _norm_text(desc_b)
        both = bool(ta) and bool(tb)
        hash_equal = both and mc_desc_hash(desc_a) == mc_desc_hash(desc_b)
        sim = _cosine(vec.vec(desc_a), vec.vec(desc_b)) if both else 0.0
        desc_agrees = hash_equal or sim >= sim_threshold
        if material_equal and desc_agrees:
            verdict = "STRONG"
        elif material_equal and not both:
            verdict = "MATERIAL_ONLY"
        elif material_equal and not desc_agrees:
            verdict = "CONFLICT"
        elif not material_equal and desc_agrees:
            verdict = "DESC_ONLY"
        else:
            verdict = "WEAK"
        return _Agreement(material_equal, hash_equal, sim, verdict)

    _WITNESS_SOURCE = "built-in fallback"
else:
    _WITNESS_SOURCE = "matchcore package"

# The description witness is now ALWAYS available (package or fallback).
_HAS_WITNESS = True


# --- per-source-type behaviour. Each As-Built source is DIFFERENT, so the
#     checks and the wording differ (Adriele: "this should be different for each
#     type"). check_batch/check_revision decide which deviations even apply. ---
SOURCE_PROFILES = {
    "Scanner":    {"check_batch": True,  "check_revision": False},  # label scanner: has batches
    "Scan":       {"check_batch": True,  "check_revision": False},  # (kept for back-compat)
    "Manual":     {"check_batch": False, "check_revision": False},
    "Reserved":   {"check_batch": False, "check_revision": False},  # Reserved Logistic
    "mb51":       {"check_batch": False, "check_revision": False},
    "Teamcenter": {"check_batch": True,  "check_revision": True},
    "As-Built":   {"check_batch": True,  "check_revision": True},
}


def profile_for(label):
    return dict(SOURCE_PROFILES.get(label,
                {"check_batch": True, "check_revision": True}))


# --- WHICH FILE IS WHICH — the As-Design vs As-Built distinction is safety-
#     critical: "In Design, not in Built" and "In Built, not in Design" mean
#     opposite things for what ships to the client. So we do NOT trust the slot
#     alone — we read each file's STRUCTURE, warn on a likely swap, and stamp the
#     detected role into the report. ---
DESIGN_MARKERS = {"Level", "TCUID", "Make/Buy", "Find Number", "Traceable",
                  "Material Type", "Occurrence Name", "Release Effectivity",
                  "Material Specification Type", "Serial Number Profile",
                  "Internal Classification", "Authorization Group"}
LIST_MARKERS = {"Material", "Charge / Batch", "Charge/Batch", "Pos", "ME",
                "Notes", "Auftrag", "Werk", "Buch.dat.", "Erfasst am", "LOrt",
                "EME", "Menge"}
ASBUILT_MARKERS = {"Serial Number", "Part Number", "Installation Time",
                   "Manufacturer's ID", "Physical UID", "Open Discrepancy",
                   "Associated Non-Conformances", "Part Used"}


def classify_source(headers):
    """Read a file's columns and say what it structurally looks like. The
    safety-critical split is a Teamcenter ENGINEERING BOM (the authoritative
    design/as-built export — rich schema: Level, TCUID, Make/Buy, Find Number,
    Serial Number, ...) versus a plain LIST / manual / SAP export (Material,
    Charge/Batch, Pos, Auftrag, Werk, ...). Telling a design MBOM apart from a
    Teamcenter As-Built by columns alone is unreliable — they share a schema —
    so we don't pretend to; we flag 'engineering BOM' vs 'list'."""
    hs = {str(h).strip() for h in headers if h is not None}
    d = len(hs & DESIGN_MARKERS)
    l = len(hs & LIST_MARKERS)
    a = len(hs & ASBUILT_MARKERS)
    tc = d + a                      # Teamcenter engineering signature
    if l >= 2 and l > tc:
        kind = "List / manual / SAP"
    elif tc >= 3:
        kind = "Teamcenter engineering BOM"
    elif l >= 1:
        kind = "List / manual / SAP"
    else:
        kind = "unknown"
    return kind, {"design": d, "list": l, "asbuilt": a}


# As-Built source types that are meant to be plain lists (NOT a Teamcenter BOM)
LIST_SOURCE_LABELS = {"Scanner", "Scan", "Manual", "Reserved", "mb51"}

# ---- macro colours (RGB) ----
COL_ASSEMBLED = "C6EFCE"  # light green
COL_MISSING   = "FFC7CE"  # light red
COL_DEVIATED  = "FFE599"  # light orange
COL_HEADER    = "D9D9D9"  # grey header band

# ---- Matched-tab block colours (Adriele: make the As-Built stand out) ----
# As-Design = blue, As-Built = orange/amber, so the two halves read as two
# distinct blocks. Strong fill on the band row, light tint on the field-header
# row; the As-Built body cells also get a faint amber wash where the row has no
# stronger status colour, and a bold divider marks where As-Built begins.
COL_DESIGN_BAND  = "2E75B6"  # blue   (white text)
COL_DESIGN_HDR   = "DDEBF7"  # light blue
COL_BUILT_BAND   = "C55A11"  # orange (white text)
COL_BUILT_HDR    = "FCE4D6"  # light orange
COL_BUILT_TINT   = "FDEDE0"  # very faint amber (As-Built body wash)


# --- SAP language-tag prefix on descriptions ---------------------------------
# Teamcenter/SAP stores the short text as "English SAP name: <desc>" (and German
# / French / ... variants). It's noise in the report and a constant token that
# adds nothing to the description match, so strip the leading "<lang> SAP name:".
_LANG_DESC = re.compile(
    r"^\s*(english|german|deutsch|french|francais|italian|italiano|spanish)"
    r"\s+sap\s+name\s*:\s*", re.I)


def _clean_desc(v):
    if v is None:
        return v
    return _LANG_DESC.sub("", str(v)).strip()

# EXACT header arrays from the macro (Merge_AsDesigned_AsBuild).
# The macro walks Array("MBOM", ...block1..., "AsBUILD", ...block2...) and, for
# each header, keeps the column only if it is FOUND in the sheet AND the column
# is not header-only/empty (HideHeaderOnlyColumns). We reproduce that exactly:
# candidate order is fixed by the macro; a column is emitted only if present and
# has at least one non-empty data cell.
MBOM_COLS_ALL = [
    "ID", "Revision", "Description", "Sequence", "Level", "Reference Designator",
    "Quantity", "Unit Of Measure", "Date Released", "Release Effectivity",
    "Owner", "Last Modified Date", "Material Specification Type", "Group ID",
    "Priority", "Plant Id", "Inspection Relevant", "Occurrence Name",
    "Internal Classification", "Authorization Group", "Serial Number Profile",
    "Serialized", "Lot", "Traceable", "Material Type", "Make/Buy", "TCUID",
    "Value", "PAD Info", "Remarks",
]
ASBUILD_COLS_ALL = [
    "ID", "Revision", "Description", "Part Used", "Usage", "Serial Number",
    "Part Number", "Installation Time", "Lot Number", "Manufacturing Date",
    "Manufacturer's ID", "Physical UID", "Open Discrepancy", "Sequence",
    "Reference Designator", "Quantity", "Unit Of Measure", "Date Released",
    "Release Effectivity", "Owner", "Last Modified Date",
    "Associated Non-Conformances", "Level", "TCUID",
]

# --- CONCISE column sets for the Matched tab (curated from the macro's list) ---
# Only the fields that matter for an As-Design vs As-Built traceability check.
# The Unmatched tabs still use the full ASBUILD/MBOM sets above.
# As-Design block, Adriele's order and MINIMAL set (she feeds the report's own
# tabs back in each round, so keep only what matters): DESCRIPTION in front of
# the material (ID), then Revision, Quantity, and the batch-relevant flags.
# Nothing else is carried.
MBOM_COLS_CONCISE = [
    "Revision Name", "Description", "ID", "Revision", "Quantity", "Traceable", "Lot",
]
# FIXED canonical As-Built columns — the ONLY As-Built fields carried anywhere
# (Matched tab and the "In <source>, not in Design" tab). _adapt_source maps
# every source onto these names; everything else is dropped.
ASBUILT_CANONICAL = [
    "Description", "Material", "Revision", "Charge / Batch", "Serial", "Qty",
    "Equipment",
]
ASBUILD_COLS_CONCISE = [
    "ID", "Revision", "Part Number", "Serial Number", "Lot Number", "Part Used",
    "Installation Time", "Manufacturing Date", "Manufacturer's ID",
    "Open Discrepancy", "Associated Non-Conformances", "Quantity",
]


# --- pattern-based part-number detection, for As-Built sources that are NOT
#     Teamcenter exports (e.g. the label-scanner CLEAN file). A material/part
#     number: 0-3 letters, 6-8 digits, optional short suffix. ---
RE_MAT = re.compile(r"^[A-Z]{0,3}\d{6,8}(?:-?[A-Z0-9]{1,3})?$")


def _looks_material(v):
    if v is None:
        return False
    return bool(RE_MAT.match(re.sub(r"[^A-Z0-9]", "", str(v).upper())))


def _pattern_material_col(headers, records):
    """Header whose values most look like part numbers (>=50%)."""
    best = (None, 0.0)
    for h in headers:
        vals = [rec.get(h) for rec in records if rec.get(h) not in (None, "")]
        if len(vals) < 3:
            continue
        frac = sum(_looks_material(v) for v in vals) / len(vals)
        if frac > best[1]:
            best = (h, frac)
    return best[0] if best[1] >= 0.5 else None


def _adapt_source(headers, records):
    """Make a non-Teamcenter As-Built source look like a Teamcenter As-Built by
    mapping its columns onto the names the comparator expects. A real Teamcenter
    export (already has 'Part Number') is left untouched. The label-scanner CLEAN
    file (Material / Charge-Batch / Serial ...) gets mapped:
        Material       -> Part Number  (+ ID)     [pattern-detected if renamed]
        Charge / Batch -> Lot Number
        Serial         -> Serial Number
    and each record is flagged _is_scan so the revision check is skipped
    (a scanned ticket carries no design revision)."""
    hs = set(headers)
    if "Part Number" in hs:
        return                                   # genuine As-Built export
    # Only adapt an ALTERNATE as-built source (the scanner CLEAN file). The
    # design MBOM has 'ID' and no 'Material' -> leave it alone. The scan has
    # 'Material' (or no 'ID' at all).
    if not ("Material" in hs or "ID" not in hs):
        return
    # Pick the MATERIAL column. Their sources call it "Material" (never "Part
    # Number"). Accept, in order: an exact 'Material'; a 'Material*Built*' column
    # (an ABCL comparison export carries BOTH a design and a built material — the
    # built side is the as-built one); any header containing 'material'; else the
    # column whose values most look like material numbers.
    def _find_material(hdrs):
        for h in hdrs:
            if str(h).strip().casefold() == "material":
                return h
        for h in hdrs:
            hl = str(h).casefold()
            if "material" in hl and "built" in hl:
                return h
        for h in hdrs:
            if "material" in str(h).casefold():
                return h
        return None
    pn_col = _find_material(headers) or _pattern_material_col(headers, records)
    lot_col = next((c for c in ("Charge / Batch", "Charge/Batch", "Batch",
                                "Lot Number", "Charge") if c in hs), None)
    # DESCRIPTION column — needed for the vectorisation witness to work. These
    # sources rarely have a column literally called 'Description': mb51/SAP call
    # it 'Materialkurztext', an ABCL export the built desc is 'Bezeichnung'. Map
    # whichever exists onto 'Description' so the engine has text to compare.
    desc_col = next((c for c in ("Description", "Bezeichnung", "Materialkurztext",
                                 "Materialkurztext (Built)", "Kurztext",
                                 "AuftragsKurztext") if c in hs), None)
    # AS-BUILT revision (prefer the built side on an ABCL export).
    rev_col = None
    for cand in ("RevStand (Built)", "Revision (Built)", "AB Rev", "RevStand",
                 "Revision", "Rev"):
        if cand in hs:
            rev_col = cand
            break
    serial_col = next((c for c in ("Serial", "SerialNr", "Serial Number",
                                   "Seriennummer", "SerNr") if c in hs), None)
    qty_col = next((c for c in ("Qty", "Quantity", "Menge", "Anzahl") if c in hs),
                   None)
    equip_col = next((c for c in ("Equipment", "Equipment Number", "Equipment No",
                                  "Equi", "Ausruestung", "Ausrüstung",
                                  "Equipmentnr", "Equipment Nr") if c in hs), None)
    if pn_col is None:
        return                                   # nothing part-number-like; leave as-is
    for rec in records:
        v = rec.get(pn_col)
        rec["Part Number"] = v
        rec.setdefault("ID", v)
        # Standardise onto the CANONICAL As-Built fields Adriele wants shown —
        # Description, Material, Revision, Charge / Batch, Serial, Qty, Equipment —
        # regardless of what the source called them. Everything else is dropped
        # from the Matched As-Built block.
        if rec.get("Material") in (None, ""):
            rec["Material"] = v
        if desc_col and rec.get("Description") in (None, ""):
            rec["Description"] = rec.get(desc_col)
        if rev_col and rec.get("Revision") in (None, ""):
            rec["Revision"] = rec.get(rev_col)
        if lot_col:
            batch_val = rec.get(lot_col)
            if rec.get("Charge / Batch") in (None, ""):
                rec["Charge / Batch"] = batch_val
            if rec.get("Lot Number") in (None, ""):
                rec["Lot Number"] = batch_val     # kept for the batch check
        if serial_col and rec.get("Serial") in (None, ""):
            rec["Serial"] = rec.get(serial_col)
        if rec.get("Serial Number") in (None, ""):
            rec["Serial Number"] = rec.get("Serial")
        if qty_col and rec.get("Qty") in (None, ""):
            rec["Qty"] = rec.get(qty_col)
        if equip_col and rec.get("Equipment") in (None, ""):
            rec["Equipment"] = rec.get(equip_col)
        rec["_is_scan"] = True


def child_key(tcuid):
    """Macro intent: strip the parent UID, key on the child token after ':'."""
    if tcuid is None:
        return ""
    s = str(tcuid).strip()
    if s == "" or s.lower() == "none":
        return ""
    return s.split(":")[-1].strip()


def _best_sheet(wb):
    """Pick the worksheet that actually holds the parts list — the one with the
    most material-number-looking values in any single column. This is essential
    for multi-sheet workbooks: the scan CLEAN file has a big 'Labels' sheet AND a
    small 'Attention' sheet, and its ACTIVE sheet is 'Attention' — so blindly
    reading wb.active compared the tiny review pile against the design and made
    almost everything look unmatched. Falls back to wb.active if nothing scores."""
    best, best_score = None, -1
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        ncol = max((len(r) for r in rows), default=0)
        score = 0
        for c in range(ncol):
            s = sum(1 for r in rows[1:] if c < len(r) and _looks_material(r[c]))
            if s > score:
                score = s
        if score > best_score:
            best_score, best = score, ws
    return best or wb.active


# Column names that mark a real header row (design MBOM, SAP list, or ABCL).
_HEADER_TOKENS = {
    "id", "material", "material (design)", "material (built)", "part number",
    "revision", "revstand", "revstand (design)", "revstand (built)", "level",
    "stufe", "description", "bezeichnung", "materialkurztext", "quantity",
    "menge", "charge", "charge / batch", "lot", "lot number", "tcuid",
    "find number", "make/buy", "traceable", "serialnr", "serial number",
    "source", "auftrag", "werk", "buch.dat.", "unit of measure",
}


def _find_header_row(rows, scan=15):
    """Index of the row that most looks like the column-header row.

    A file may have a note banner or blank rows above the headers (e.g. an ADAB
    'In Design, not in ...' tab exported as-is). Score the first `scan` rows by
    how many cells match a known header name; the best (>=2 hits) wins. Falls
    back to the first non-empty row so behaviour is unchanged for clean files."""
    best_i, best_score = None, 0
    for i, r in enumerate(rows[:scan]):
        score = 0
        for c in r:
            if c is None:
                continue
            if str(c).strip().casefold() in _HEADER_TOKENS:
                score += 1
        if score > best_score:
            best_score, best_i = score, i
    if best_i is not None and best_score >= 2:
        return best_i
    for i, r in enumerate(rows):          # fallback: first non-empty row
        if any(c is not None and str(c).strip() != "" for c in r):
            return i
    return 0


def load_bom(path):
    """Return (headers, list-of-row-dicts) from the parts-list sheet, row1=header.

    Teamcenter exports write a wrong worksheet <dimension> tag (often ref="A1"),
    which makes openpyxl's fast read_only reader stop after the first cell and
    report the sheet as empty. The normal reader ignores the dimension and scans
    every <row> element, so we use it unconditionally here. On a per-unit BOM
    (tens to low-hundreds of rows) the speed difference is irrelevant.

    Reads the sheet with the most material-looking rows (not blindly wb.active),
    so a CLEAN file whose active sheet is 'Attention' is still read from 'Labels'.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = _best_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    # Find the REAL header row. A file exported from a previous ADAB run (or any
    # SAP/Excel export) often has a NOTE banner or blank line above the column
    # names — e.g. the "In Design, not in ..." tab keeps its note in row 1 and the
    # real headers (Level, ID, Revision, ...) in row 2. Reading row 1 as headers
    # then finds no ID/Material column and the whole side reads as ZERO parts, so
    # nothing matches. Scan the first rows and pick the one that looks most like a
    # header (most known BOM/list column names).
    hi = _find_header_row(rows)
    header = [h for h in rows[hi]]
    # first occurrence wins for duplicate headers (e.g. two 'Level' cols)
    idx = {}
    for i, h in enumerate(header):
        if h is not None and h not in idx:
            idx[h] = i
    records = []
    for r in rows[hi + 1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rec = {h: (r[i] if i < len(r) else None) for h, i in idx.items()}
        rec["_key"] = child_key(rec.get("TCUID"))
        # strip the "English SAP name:" language tag off the description
        if "Description" in rec:
            rec["Description"] = _clean_desc(rec["Description"])
        # Revision Name is the SAP short text (same language as SAP/mb51) — clean
        # it the same way so it can be matched/vectorised as the reliable "code".
        if "Revision Name" in rec:
            rec["Revision Name"] = _clean_desc(rec["Revision Name"])
        records.append(rec)
    # adapt a non-Teamcenter As-Built source (e.g. the scanner CLEAN file) so it
    # speaks the same column names as a Teamcenter As-Built export
    _adapt_source(list(idx.keys()), records)
    return list(idx.keys()), records


def get(rec, col):
    v = rec.get(col)
    return "" if v is None else v


# Columns whose values are genuinely NUMERIC and should be written to Excel as
# numbers, not text — otherwise Excel shows the green "number stored as text"
# triangle and SUM/filter misbehave. Part numbers, revisions, batches and
# serials are deliberately NOT here: they are identifiers and must stay text
# (e.g. keep leading zeros / the exact string).
_NUMERIC_COLS = {"Quantity", "Qty", "Menge", "Design Qty", "Built Qty"}


def _num(v):
    """Return v as int/float if it is a number written as text, else unchanged."""
    if isinstance(v, bool) or isinstance(v, (int, float)):
        return v
    if v is None:
        return v
    s = str(v).strip()
    if s == "":
        return v
    t = s.replace(",", ".") if ("," in s and "." not in s) else s
    try:
        f = float(t)
    except (ValueError, TypeError):
        return v
    return int(f) if f.is_integer() else f


def cell_val(col, v):
    """Value to write for column `col`: coerce known numeric columns to numbers."""
    return _num(v) if col in _NUMERIC_COLS else v


def present_nonempty(candidate_cols, headers, records):
    """Macro rule: keep a candidate header only if it exists in the sheet AND
    the column has at least one non-empty data cell (HideHeaderOnlyColumns)."""
    header_set = set(headers)
    out = []
    for col in candidate_cols:
        if col not in header_set:
            continue
        has_data = any(
            rec.get(col) is not None and str(rec.get(col)).strip() != ""
            for rec in records
        )
        if has_data:
            out.append(col)
    return out


def actual_cols(headers, records):
    """Every real column present in the source file that has data (the WHOLE
    row), excluding internal helper keys. Used for the Unmatched tabs so they
    show the entire report — all of the scan's columns (Picture, Type, Material,
    Charge/Batch, Serial, Order, …), not a curated subset."""
    out = []
    for h in headers:
        if h is None or str(h).startswith("_"):
            continue
        if any(rec.get(h) not in (None, "") for rec in records):
            out.append(h)
    return out


def _norm_key(v):
    """Normalise a match key — case-, type- and whitespace-invariant, but
    suffix-preserving (Adriele's rule: 7004369 == 7004369.0, case never matters,
    but C3529115-C != C3529115). Uses matchcore.norm_key when available; the
    fallback below is byte-identical in behaviour so results never depend on
    whether the package is present."""
    if _HAS_MATCHCORE:
        return mc_norm_key(v)
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r"\s+", " ", str(v).strip().upper())


def design_key(rec):
    """Design match key — the engineering part number in the design BOM.
    Prefers 'ID'; falls back to Part Number / Material so ANY parts list can be
    used as the As-Design side, not only a Teamcenter MBOM."""
    for col in ("ID", "Part Number", "Material"):
        v = rec.get(col)
        if v not in (None, ""):
            return _norm_key(v)
    return ""


def built_key(rec):
    """Built match key — the part number on the built/list side. Prefers
    'Part Number', then 'Material', then 'ID'. The fallback matters: a design-
    style file (only an 'ID' column) or a manual list can now be used as the
    As-Built side too, instead of reading as ZERO parts."""
    for col in ("Part Number", "Material", "ID"):
        v = rec.get(col)
        if v not in (None, ""):
            return _norm_key(v)
    return ""


def _to_float(v):
    try:
        return float(str(v).strip())
    except (ValueError, AttributeError):
        return None


def evaluate_pair(design_rec, built_rec, design_qty_by_id, built_qty_by_pn,
                  profile=None):
    """Return (status_text, severity) for a matched pair.

    The checks that apply depend on the SOURCE profile (Adriele: different per
    type). A Manual / Reserved / mb51 list is just a list of part numbers — it
    is NOT expected to carry batches or revisions, so those checks are OFF and a
    present part is simply 'Assembled' (green). A Scan or Teamcenter As-Built DOES
    carry batches, so a traceable design part with no batch is still flagged RED.

    severity: 'red' > 'orange' > 'green'. Multiple issues listed; colour = worst.
    """
    if profile is None:
        profile = {"check_batch": True, "check_revision": True}
    issues = []
    severity = "green"

    # --- batch / traceability (RED) --- only when this source is expected to
    # carry batches (check_batch). A manual list never is -> no false flag.
    if profile.get("check_batch", True):
        lot = str(get(built_rec, "Lot Number")).strip()
        traceable = str(get(design_rec, "Traceable")).strip().casefold()
        is_lot = str(get(design_rec, "Lot")).strip().casefold()
        needs_batch = traceable == "true" or is_lot == "true"
        if needs_batch and (lot == "" or lot == "-"):
            issues.append("No batch/traceability")
            severity = "red"

    # --- quantity (ORANGE) ---
    k = design_key(design_rec)
    dq = design_qty_by_id.get(k)
    bq = built_qty_by_pn.get(k)
    if dq is not None and bq is not None and abs(dq - bq) > 1e-9:
        if bq > dq:
            issues.append(f"Quantity over (design {dq:g}, built {bq:g})")
        else:
            issues.append(f"Quantity short (design {dq:g}, built {bq:g})")
        if severity != "red":
            severity = "orange"

    # --- revision (ORANGE) --- only when this source carries a design revision
    # (check_revision) and it is not a scan ticket.
    if profile.get("check_revision", True) and not built_rec.get("_is_scan"):
        drev = str(get(design_rec, "Revision")).strip()
        brev = str(get(built_rec, "Revision")).strip()
        if drev != "" and brev == "":
            issues.append("Built revision missing")
            if severity != "red":
                severity = "orange"
        elif drev != "" and brev != "" and drev.casefold() != brev.casefold():
            issues.append(f"Revision changed ({drev} -> {brev})")
            if severity != "red":
                severity = "orange"

    if not issues:
        return "Assembled", "green"
    return "Deviated - " + "; ".join(issues), severity


def severity_fill(severity):
    if severity == "green":
        return PatternFill("solid", fgColor=COL_ASSEMBLED)
    if severity == "orange":
        return PatternFill("solid", fgColor=COL_DEVIATED)
    if severity == "red":
        return PatternFill("solid", fgColor=COL_MISSING)
    return None


def status_fill(status):
    s = status.casefold()
    if s == "assembled":
        return PatternFill("solid", fgColor=COL_ASSEMBLED)
    if "missing" in s:
        return PatternFill("solid", fgColor=COL_MISSING)
    if "deviated" in s:
        return PatternFill("solid", fgColor=COL_DEVIATED)
    return None


def _is_level0(rec):
    """True if this is the top-level assembly row (Level == 0)."""
    v = rec.get("Level")
    if v is None:
        return False
    s = str(v).strip()
    return s == "0" or s == "0.0"


def split_three(design_recs, built_recs):
    """Match built.Part Number <-> design.ID, quantity/position aware.

    A design ID may appear on several design lines (same part used in more than
    one position) and/or the unit may install several copies of one part.
    Matching model (DESIGN-DRIVEN — every design position is kept):
      - Each DESIGN position whose part number was built -> Matched. Built copies
        of that part are cycled across its positions so batches/serials vary.
      - If a part was built in MORE copies than it has design positions, the
        surplus copies are also emitted as Matched (higher-than-designed
        quantity), paired to the last design position of that part.
      - Design positions of a part that was NEVER built -> Unmatched AsDesign.
      - Built copies whose part number is NOT any design ID -> Unmatched AsBuilt.
    Conservation is asserted at the end: no design position and no built copy is
    ever silently dropped (the earlier built-driven code lost surplus design
    positions of matched parts — roughly half the lines on a real MBOM).

    The top-level assembly row (Level 0) is excluded from all three groups: it
    is the assembly itself, not a component, and comparing its revision to the
    built unit revision produces a spurious deviation.

    Returns: (matched, unmatched_built, unmatched_design)
      matched          list of (design_rec, built_rec)
      unmatched_built  list of built_rec
      unmatched_design list of design_rec
    """
    design_recs = [d for d in design_recs if not _is_level0(d)]
    built_recs = [b for b in built_recs if not _is_level0(b)]

    # pools keyed by normalised part number
    design_pool = {}                     # ID -> list of design line records
    for d in design_recs:
        k = design_key(d)
        if k:
            design_pool.setdefault(k, []).append(d)
    built_pool = {}                      # part number -> list of built copies
    for b in built_recs:
        k = built_key(b)
        if k:
            built_pool.setdefault(k, []).append(b)

    matched = []

    # DESIGN-DRIVEN matching (fixes the dropped-lines bug).
    # The OLD code iterated the BUILT copies to make matched pairs, so a part
    # designed on MANY positions but built in FEWER copies lost its extra design
    # positions from the report entirely: they were never matched, and — because
    # the part WAS built (consumed > 0) — never listed as "In Design, not in ...".
    # Result: a report with ~half the input lines. Now we iterate the DESIGN
    # positions instead, so EVERY design line whose part was built becomes a
    # matched row. Built copies are cycled across a part's design positions so
    # their batches/serials still vary down the rows.
    used = {}
    for d in design_recs:
        k = design_key(d)
        if not k or k not in built_pool:
            continue
        copies = built_pool[k]
        i = used.get(k, 0)
        b = copies[i] if i < len(copies) else copies[-1]
        used[k] = i + 1
        matched.append((d, b))

    # If a part was built in MORE copies than it has design positions, the extra
    # built copies (a higher-than-designed quantity) are still real As-Built
    # evidence — append them paired to the LAST design position of that part, so
    # no built copy is dropped from the report either.
    for k, copies in built_pool.items():
        if k in design_pool:
            shown = used.get(k, 0)
            if shown < len(copies):
                last_d = design_pool[k][-1]
                for b in copies[shown:]:
                    matched.append((last_d, b))

    # Unmatched As-Design = design positions whose part was NEVER built.
    unmatched_design = [d for d in design_recs
                        if design_key(d) and design_key(d) not in built_pool]

    # Unmatched As-Built = built copies whose part is not in the design at all.
    unmatched_built = [b for b in built_recs
                       if not built_key(b) or built_key(b) not in design_pool]

    # --- CONSERVATION GUARANTEE (Adriele: "we cannot exclude lines") ----------
    # Every non-level-0 design position with a part number must appear exactly
    # once across matched + unmatched_design; every built copy with a part number
    # must appear at least once across matched + unmatched_built. Verify it here
    # so the report can NEVER silently shrink again.
    _d_keyed = {id(d) for d in design_recs if design_key(d)}
    _d_out = {id(d) for d, _ in matched} | {id(d) for d in unmatched_design}
    _b_keyed = {id(b) for b in built_recs if built_key(b)}
    _b_out = {id(b) for _, b in matched} | {id(b) for b in unmatched_built}
    if _d_keyed != _d_out or not _b_keyed <= _b_out:
        raise RuntimeError(
            "ADAB conservation check FAILED — some lines would be dropped "
            f"(design in {len(_d_keyed)}, out {len(_d_out)}; "
            f"built in {len(_b_keyed)}, shown {len(_b_out & _b_keyed)}). "
            "This is a bug; report it rather than trusting the output.")

    return matched, unmatched_built, unmatched_design


def part_reconciliation(design_recs, built_recs, label="As-Built"):
    """DISTINCT-PART reconciliation — the auditable headline numbers.

    `label` names the As-Built source so the MISSING wording is right per type
    ("MISSING (not in Manual)", not "not built"). A Fingerprint column carries
    matchcore's stable hash of material+description — identical content -> same
    hash across any list, so you can spot the same part regardless of source.

    Both sides are grouped by part number and compared on SUMMED quantity, so
    the result is fully reproducible and independent of how many BOM occurrences
    (positions) a part has in the design tree. This is the number a reviewer or
    auditor can recompute by hand; the position-level Matched / Unmatched tabs
    hold the detail behind it.

    A part is:
      MATCHED  in both design and built (SHORT/OVER if summed qty differs)
      MISSING  in design, never built
      EXTRA    built, not in the design baseline at all
    Returns (rows, counts).
    """
    design_recs = [d for d in design_recs if not _is_level0(d)]
    built_recs = [b for b in built_recs if not _is_level0(b)]

    dz = {}
    for d in design_recs:
        k = design_key(d)
        if not k:
            continue
        q = _to_float(d.get("Quantity"))
        e = dz.setdefault(k, {"qty": 0.0, "pos": 0, "rec": d, "hasqty": False})
        if q is not None:
            e["qty"] += q
            e["hasqty"] = True
        e["pos"] += 1

    bz = {}
    for b in built_recs:
        k = built_key(b)
        if not k:
            continue
        q = _to_float(b.get("Quantity"))
        e = bz.setdefault(k, {"qty": 0.0, "copies": 0, "rec": b, "hasqty": False})
        if q is not None:
            e["qty"] += q
            e["hasqty"] = True
        e["copies"] += 1

    counts = {"design_parts": len(dz), "built_parts": len(bz),
              "matched": 0, "missing": 0, "extra": 0, "short": 0, "over": 0}
    rows = []
    for k in sorted(set(dz) | set(bz)):
        d, b = dz.get(k), bz.get(k)
        dq = d["qty"] if (d and d["hasqty"]) else None
        bq = b["qty"] if (b and b["hasqty"]) else None
        # Fingerprint text = the SAP short name. On the design side that is the
        # Revision Name (matches SAP/mb51 exactly); Description is English and
        # would not. This makes the material+text "code" comparable across sources.
        if d:
            desc = get(d["rec"], "Revision Name") or get(d["rec"], "Description")
        elif b:
            desc = get(b["rec"], "Description")
        else:
            desc = ""
        if d and b:
            counts["matched"] += 1
            if dq is not None and bq is not None and abs(dq - bq) > 1e-9:
                status = "SHORT" if bq < dq else "OVER"
                counts["short" if bq < dq else "over"] += 1
            else:
                status = "MATCHED"
        elif d and not b:
            counts["missing"] += 1
            status = f"MISSING (not in {label})"
        else:
            counts["extra"] += 1
            status = "EXTRA (not in design)"
        fp = mc_record_hash(k, desc)[:12]
        rows.append({
            "Part Number": k, "Description": desc, "Status": status,
            "In Design": "Y" if d else "", "In Built": "Y" if b else "",
            "Design Qty": (round(dq, 4) if dq is not None else ""),
            "Built Qty": (round(bq, 4) if bq is not None else ""),
            "Design Positions": (d["pos"] if d else 0),
            "Built Copies": (b["copies"] if b else 0),
            "Fingerprint": fp,
        })
    return rows, counts


def write_summary_tab(wb, counts, unit_name, label="As-Built", meta=None):
    """Auditable headline sheet — distinct-part counts that reconcile by hand,
    STAMPED with which file was the As-Design (authority) and which was the
    As-Built, and what each was structurally detected as."""
    ws = wb.create_sheet(title="Summary")
    big = Font(name="Arial", size=13, bold=True)
    hf = Font(name="Arial", size=11, bold=True)
    sm = Font(name="Arial", size=10)
    ws.cell(row=1, column=1, value=f"ADAB reconciliation - {unit_name}").font = big

    # --- ROLE STAMP: unambiguous record of who is As-Design vs As-Built ---
    row = 2
    if meta:
        ds = ws.cell(row=row, column=1,
                     value=f"AS-DESIGN (authority):  {meta.get('design_name') or '(design)'}"
                           f"   [detected: {meta.get('design_kind','?')}]")
        ds.font = hf
        ds.fill = PatternFill("solid", fgColor="D9E1F2")
        row += 1
        bs = ws.cell(row=row, column=1,
                     value=f"AS-BUILT ({label}):  {meta.get('built_name')}"
                           f"   [detected: {meta.get('built_kind','?')}]")
        bs.font = hf
        bs.fill = PatternFill("solid", fgColor="FCE4D6")
        row += 1
        if meta.get("swap_warn"):
            w = ws.cell(row=row, column=1,
                        value=f"!!! CHECK ROLES: {meta['swap_warn']}")
            w.font = Font(name="Arial", size=11, bold=True, color="9C0006")
            w.fill = PatternFill("solid", fgColor="FFC7CE")
            row += 1
        row += 1
    ws.cell(row=row, column=1,
            value="Distinct part numbers. Each count is a set operation on part "
                  "numbers + summed quantity, so it can be reproduced by hand.").font = sm
    _summary_start = row + 2

    lines = [
        ("Design - distinct parts", counts["design_parts"], None),
        ("Built - distinct parts", counts["built_parts"], None),
        ("", "", None),
        ("Matched (in both)", counts["matched"], COL_ASSEMBLED),
        ("  of which quantity SHORT", counts["short"], None),
        ("  of which quantity OVER", counts["over"], None),
        (f"MISSING - in design, not in {label}", counts["missing"], COL_MISSING),
        (f"EXTRA - in {label}, not in design", counts["extra"], COL_DEVIATED),
    ]
    r = _summary_start
    for text, val, colour in lines:
        cl = ws.cell(row=r, column=1, value=text)
        cl.font = sm if text.startswith("  ") else hf
        ws.cell(row=r, column=2, value=val)
        if colour:
            for cc in (1, 2):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=colour)
        r += 1
    ws.cell(row=r + 1, column=1,
            value=f"Identity check: matched {counts['matched']} + missing "
                  f"{counts['missing']} = {counts['matched'] + counts['missing']} "
                  f"distinct design parts.").font = sm
    ws.column_dimensions["A"].width = 66
    ws.column_dimensions["B"].width = 14


def _status_rank(status):
    """Problems first, MATCHED last — independent of the per-type wording."""
    if status.startswith("MISSING"):
        return 0
    if status.startswith("EXTRA"):
        return 1
    if status == "SHORT":
        return 2
    if status == "OVER":
        return 3
    return 4


def write_reconciliation_tab(wb, rows, key_label="Part Number"):
    """One row per DISTINCT part: design vs built qty, positions, status, hash.
    `key_label` names the key column ("Material" for a scan/SAP/mb51/ABCL list,
    "Part Number" for a Teamcenter As-Built) — the values are unchanged, only the
    header word."""
    ws = wb.create_sheet(title="Parts Reconciliation")
    cols = ["Part Number", "Description", "Status", "In Design", "In Built",
            "Design Qty", "Built Qty", "Design Positions", "Built Copies",
            "Fingerprint"]
    display = [key_label if c == "Part Number" else c for c in cols]
    hf = Font(name="Arial", size=9, bold=True)
    hfill = PatternFill("solid", fgColor=COL_HEADER)
    thin = Font(name="Arial", size=9)
    for c, h in enumerate(display, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
    for i, row in enumerate(
            sorted(rows, key=lambda x: (_status_rank(x["Status"]), x["Part Number"])),
            start=2):
        for c, col in enumerate(cols, start=1):
            ws.cell(row=i, column=c, value=row.get(col, "")).font = thin
        st = row["Status"]
        fill = None
        if st.startswith("MISSING"):
            fill = PatternFill("solid", fgColor=COL_MISSING)
        elif st.startswith("EXTRA") or st in ("SHORT", "OVER"):
            fill = PatternFill("solid", fgColor=COL_DEVIATED)
        elif st == "MATCHED":
            fill = PatternFill("solid", fgColor=COL_ASSEMBLED)
        if fill:
            for c in range(1, len(cols) + 1):
                ws.cell(row=i, column=c).fill = fill
    ws.freeze_panes = "A2"
    for c, w in enumerate([16, 34, 22, 9, 9, 11, 11, 15, 13, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def sheet_name_for(path):
    base = os.path.basename(path)
    m = re.search(r"(CH\d{3,5})", base)
    if m:
        return m.group(1)
    return os.path.splitext(base)[0][:31]


def _safe_name(stem):
    """Make a string safe to use as a Windows filename: drop the characters
    Windows forbids (\\ / : * ? \" < > |) and trim length."""
    stem = re.sub(r'[\\/:*?"<>|]+', "", stem)
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem[:180] or "adab_report"


def _band(ws, row, col, text, hdr_font, hdr_fill):
    c = ws.cell(row=row, column=col, value=text)
    c.font = hdr_font
    c.fill = hdr_fill
    return c


def _engine_label(ag):
    """Plain-words statement of WHICH witness confirmed a matched row. Every
    matched row already agrees on the Material number (that's how it matched);
    this says whether the DESCRIPTION vectorisation agreed too."""
    if ag.verdict == "STRONG":
        return "Material + Description"
    if ag.verdict == "MATERIAL_ONLY":
        return "Material only (no description)"
    if ag.verdict == "CONFLICT":
        return "Material (DESC conflict!)"
    return "Material"


def _name_label(ag):
    """Plain-words verdict for the REVISION NAME (SAP short text) check: is the
    design's Revision Name the same as the As-Built short text? This is the
    reliable one for SAP sources — Revision Name is stored in SAP's language,
    unlike the English 'Description'."""
    if ag.desc_hash_equal or ag.desc_similarity >= 0.999:
        return "Same name"
    if ag.desc_similarity >= 0.60:
        return "Similar name"
    if not str(ag.desc_similarity):
        return "-"
    return "Name differs"


def write_description_rescue_tab(wb, unmatched_built, unmatched_design,
                                 vectorizer, key_label="Material"):
    """THE DESCRIPTION ENGINE, made visible. Take every EXTRA (in built, not in
    design by material) and every MISSING (in design, not in built by material),
    transform each description with the TF-IDF vectoriser and compare — so a part
    whose NUMBER differs (typo, relabel, OCR slip, different source coding) but
    whose DESCRIPTION is the same surfaces as a rescue candidate instead of a
    false extra/missing. Sorted best-similarity first. This is exactly 'take the
    description, transform it and compare it as if it were a material number'."""
    ws = wb.create_sheet(title="Review - Name match")
    hf = Font(name="Arial", size=9, bold=True)
    thin = Font(name="Arial", size=9)
    hfill = PatternFill("solid", fgColor=COL_HEADER)
    cols = [f"Built {key_label}", "Built Short Text",
            "Design ID", "Design Rev Name",
            "Name Similarity", "Verdict"]
    note = ("Same part, DIFFERENT number? These pairs did NOT match on the "
            f"{key_label} number, but the NAME engine (TF-IDF vectorise -> cosine "
            "on the Revision Name / SAP short text) found them similar. Review "
            "each: a real match here means one EXTRA and one MISSING above are "
            "actually the same part.")
    nb = ws.cell(row=1, column=1, value=note)
    nb.font = hf
    nb.fill = PatternFill("solid", fgColor=COL_DEVIATED)
    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hf
        cell.fill = hfill

    # index the missing (design) Revision Names once (SAP short text — the field
    # that matches SAP/mb51, unlike the English Description).
    miss = [(design_key(d), get(d, "Revision Name") or get(d, "Description"))
            for d in unmatched_design]
    miss = [(k, dsc) for k, dsc in miss if k and str(dsc).strip()]
    rows = []
    for b in unmatched_built:
        bk = built_key(b)
        bdesc = get(b, "Description")
        if not bk or not str(bdesc).strip():
            continue
        bvec = vectorizer.vec(bdesc)
        best = None
        for mk, mdesc in miss:
            ag = mc_agree(bk, bdesc, mk, mdesc, vectorizer)
            if ag.verdict == "DESC_ONLY" or ag.desc_similarity >= 0.60:
                if best is None or ag.desc_similarity > best[4]:
                    best = [bk, bdesc, mk, mdesc,
                            round(ag.desc_similarity, 3), ag.verdict]
        if best:
            rows.append(best)
    # de-duplicate: a combined list repeats the same material on many lines, so
    # collapse identical (built material -> design ID) candidate pairs.
    seen, uniq = set(), []
    for r in rows:
        sig = (mc_norm_key(r[0]), mc_norm_key(r[2]))
        if sig in seen:
            continue
        seen.add(sig)
        uniq.append(r)
    rows = uniq
    rows.sort(key=lambda r: r[4], reverse=True)
    for i, r in enumerate(rows, start=3):
        for c, v in enumerate(r, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = thin
            cell.fill = PatternFill("solid", fgColor=COL_DEVIATED)
    ws.freeze_panes = "A3"
    for c, w in enumerate([16, 34, 16, 34, 14, 12], start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return len(rows)


def write_matched_tab(wb, matched, mbom_cols, asb_cols,
                      design_qty_by_id, built_qty_by_pn,
                      profile=None, vectorizer=None):
    ws = wb.create_sheet(title="Matched")
    thin = Font(name="Arial", size=9)
    hf = Font(name="Arial", size=9, bold=True)
    hfill = PatternFill("solid", fgColor=COL_HEADER)

    # description-witness columns. ALWAYS on now (package or built-in fallback),
    # so the report always shows WHICH ENGINE matched each row (Adriele's ask):
    #   Match Engine = plain words: "Material + Description" / "Material only" /
    #                  "Material (DESC conflict)"
    #   Desc Match   = the raw verdict (STRONG / MATERIAL_ONLY / CONFLICT ...)
    #   Desc Sim     = TF-IDF cosine of the two descriptions (the vectorisation)
    mc_cols = (["Name Match", "Name Sim", "Desc Match", "Desc Sim", "Fingerprint"]
               if vectorizer else [])

    # block colours (Adriele: make the As-Built stand out)
    white_bold = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    band_design = PatternFill("solid", fgColor=COL_DESIGN_BAND)
    band_built = PatternFill("solid", fgColor=COL_BUILT_BAND)
    fh_design = PatternFill("solid", fgColor=COL_DESIGN_HDR)
    fh_built = PatternFill("solid", fgColor=COL_BUILT_HDR)
    built_tint = PatternFill("solid", fgColor=COL_BUILT_TINT)

    n_m = len(mbom_cols)
    n_a = len(asb_cols)
    built_c0 = 2 + n_m                       # first As-Built column
    built_c1 = built_c0 + n_a - 1            # last  As-Built column
    witness_c0 = 2 + n_m + n_a               # first witness column

    _band(ws, 1, 1, "Status", hf, hfill)
    if n_m:
        _band(ws, 1, 2, "As-Design (MBOM)", white_bold, band_design)
    if asb_cols:
        _band(ws, 1, built_c0, "As-Built", white_bold, band_built)
    if mc_cols:
        _band(ws, 1, witness_c0, "Description witness", hf, hfill)

    # suffix headers that appear on both sides so they are unambiguous
    common = set(mbom_cols) & set(asb_cols)
    mbom_hdr = [f"{c} (Design)" if c in common else c for c in mbom_cols]
    asb_hdr = [f"{c} (Built)" if c in common else c for c in asb_cols]
    field_headers = ["Status"] + mbom_hdr + asb_hdr + mc_cols
    for c, h in enumerate(field_headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hf
        if 2 <= c < built_c0:
            cell.fill = fh_design
        elif built_c0 <= c <= built_c1:
            cell.fill = fh_built
        else:
            cell.fill = hfill
        cell.alignment = Alignment(horizontal="left")

    deviations = 0
    for i, (drec, brec) in enumerate(matched, start=3):
        status, severity = evaluate_pair(drec, brec,
                                         design_qty_by_id, built_qty_by_pn,
                                         profile=profile)
        if severity != "green":
            deviations += 1
        vals = [status] + [cell_val(c, get(drec, c)) for c in mbom_cols] + \
               [cell_val(c, get(brec, c)) for c in asb_cols]
        if mc_cols:
            drev = get(drec, "Revision Name")
            ddesc = get(drec, "Description")
            btext = get(brec, "Description")     # As-Built SAP short text (Materialkurztext)
            ag_name = mc_agree(design_key(drec), drev, built_key(brec), btext, vectorizer)
            ag_desc = mc_agree(design_key(drec), ddesc, built_key(brec), btext, vectorizer)
            # DB "code" = material number + Revision Name (SAP short text) — the
            # field that actually matches across design and SAP/mb51.
            fp = mc_record_hash(built_key(brec), drev)[:12]
            vals += [_name_label(ag_name), round(ag_name.desc_similarity, 3),
                     _engine_label(ag_desc), round(ag_desc.desc_similarity, 3), fp]
        fill = severity_fill(severity)
        # Highlight when the SAP name (Revision Name) does NOT agree with the
        # As-Built short text even though the material number matched.
        conflict = (bool(mc_cols) and ag_name.material_equal
                    and ag_name.desc_similarity < 0.60 and severity == "green")
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = thin
            if conflict:
                cell.fill = PatternFill("solid", fgColor=COL_DEVIATED)
            elif fill:
                cell.fill = fill
            # As-Built body wash only when the row is a plain green "Assembled"
            # (so the amber tint marks the As-Built block without hiding an
            # orange/red status).
            if built_c0 <= c <= built_c1 and not conflict and severity == "green":
                cell.fill = built_tint

    # bold divider marking where the As-Built block begins (and where it ends /
    # the witness block starts), across header + data rows.
    med = Side(style="medium", color="7F7F7F")
    last = 2 + len(matched)
    for r in range(1, last + 1):
        ws.cell(row=r, column=built_c0).border = Border(left=med)
        if mc_cols:
            ws.cell(row=r, column=witness_c0).border = Border(left=med)

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 42
    for c in range(2, len(field_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    return deviations


def write_single_side_tab(wb, title, records, cols, note, note_fill):
    ws = wb.create_sheet(title=title[:31])
    thin = Font(name="Arial", size=9)
    hf = Font(name="Arial", size=9, bold=True)
    hfill = PatternFill("solid", fgColor=COL_HEADER)

    # note banner row 1
    nb = ws.cell(row=1, column=1, value=note)
    nb.font = Font(name="Arial", size=9, bold=True)
    nb.fill = PatternFill("solid", fgColor=note_fill)

    for c, h in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="left")

    for i, rec in enumerate(records, start=3):
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=c, value=cell_val(col, get(rec, col)))
            cell.font = thin
            cell.fill = PatternFill("solid", fgColor=note_fill)

    ws.freeze_panes = "A3"
    for c in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20


def resolve_design(design_arg):
    if os.path.isdir(design_arg):
        files = [f for f in glob.glob(os.path.join(design_arg, "*.xls*"))
                 if not os.path.basename(f).startswith("~$")]
        if len(files) != 1:
            raise SystemExit(f"Design folder must hold exactly one file, found {len(files)}")
        return files[0]
    return design_arg


def _sheet_safe(text, limit=31):
    """Excel sheet name: <=31 chars, none of : \\ / ? * [ ]."""
    s = re.sub(r'[:\\/?*\[\]]+', " ", str(text)).strip()
    return (s[:limit] or "Sheet")


def _one_report(out_path, name, built_headers, built_recs,
                design_recs, design_headers, design_recs_noL0,
                mbom_cols_concise, design_qty_by_id, log,
                built_label="As-Built", design_name=""):
    """Build ONE ADAB report for a set of built records against the design.

    built_label names the As-Built source (e.g. 'Scan', 'Manual', 'Reserved',
    'mb51', 'Teamcenter'). The two unmatched tabs are named FOR the two lists
    being compared — "In Design, not in <label>" and "In <label>, not in Design"
    — so a traceability reviewer reads the direction straight off the tab.
    Returns a result dict."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if not built_recs:
        ws = wb.create_sheet(title=name[:31])
        ws.cell(row=1, column=1,
                value="EMPTY EXPORT - no As-Built data "
                      "(work orders not closed / Teamcenter export failed)")
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor=COL_MISSING)
        try:
            wb.save(out_path)
        except PermissionError:
            log(f"  COULD NOT SAVE {os.path.basename(out_path)} — it's open "
                f"in Excel. Close it and run again.")
            return {"unit": name, "error": "report open in Excel"}
        log(f"  {name}: EMPTY export -> {os.path.basename(out_path)}")
        return {"unit": name, "empty": True}

    built_recs_noL0 = [b for b in built_recs if not _is_level0(b)]
    asb_cols_concise = present_nonempty(ASBUILD_COLS_CONCISE, built_headers,
                                        built_recs_noL0)
    # scan / DocuBOM / list source (mb51, Reserved, Manual) -> show its own
    # columns in the Matched tab. The BATCH must sit right after the As-Built
    # Material (Adriele's ask) so a reviewer reads "which part / which batch" in
    # one glance. The batch column can be named several ways across sources
    # (Charge / Charge / Batch / Batch / Lot Number); take whichever exists and
    # place it immediately after Material.
    if any(b.get("_is_scan") for b in built_recs_noL0):
        # FIXED As-Built block (Adriele): always exactly these canonical fields,
        # in this order, and nothing else. _adapt_source has already standardised
        # every source's columns onto these names, so this list never changes.
        asb_cols_concise = list(ASBUILT_CANONICAL)

    built_qty_by_pn = {}
    for b in built_recs_noL0:
        k = built_key(b)
        q = _to_float(b.get("Quantity"))
        if k and q is not None:
            built_qty_by_pn[k] = built_qty_by_pn.get(k, 0.0) + q

    matched, unmatched_built, unmatched_design = split_three(
        design_recs, built_recs)

    # Line-conservation report (Adriele: "we cannot exclude lines"). Every design
    # position must land in Matched or "In Design, not in ..."; every built copy
    # in Matched or "In ..., not in Design". Show the user the sums balance.
    n_design_lines = len([d for d in design_recs
                          if not _is_level0(d) and design_key(d)])
    n_built_lines = len([b for b in built_recs
                         if not _is_level0(b) and built_key(b)])
    n_matched_design = len({id(d) for d, _ in matched})
    log(f"  Line conservation -> DESIGN: {n_matched_design} matched-positions + "
        f"{len(unmatched_design)} in-design-not-{built_label} = "
        f"{n_matched_design + len(unmatched_design)} of {n_design_lines} input "
        f"design lines  |  BUILT: {len(matched)} matched-rows + "
        f"{len(unmatched_built)} in-{built_label}-not-design "
        f"(from {n_built_lines} built lines).")

    # --- role check: is the right file in the right slot? ---
    design_kind, dsig = classify_source(design_headers)
    built_kind, bsig = classify_source(built_headers)
    swap_warn = None
    if design_kind.startswith("List"):
        swap_warn = ("the As-Design looks like a plain list / SAP export, NOT a "
                     "Teamcenter engineering baseline. Put the AUTHORITATIVE "
                     "design file in the As-Design slot.")
    elif built_kind.startswith("Teamcenter") and built_label in LIST_SOURCE_LABELS:
        swap_warn = (f"the As-Built is a Teamcenter engineering BOM, but you "
                     f"marked it as '{built_label}' (a plain list). If that file "
                     f"is really the design, you swapped As-Design and As-Built.")
    if swap_warn:
        log(f"  !!! ROLE WARNING: {swap_warn}")
    log(f"  Roles -> As-Design: {design_name or '(design)'} [{design_kind}]  |  "
        f"As-Built ({built_label}): {name} [{built_kind}]")
    meta = {"design_name": design_name, "design_kind": design_kind,
            "built_name": name, "built_kind": built_kind,
            "built_label": built_label, "swap_warn": swap_warn}

    # per-source-type behaviour (which checks apply, wording)
    profile = profile_for(built_label)
    # If the source carries NO batch/lot data at all, never flag a batch gap —
    # a guard on top of the profile so a mislabelled source can't false-flag.
    if not any(str(get(b, "Lot Number")).strip() for b in built_recs_noL0):
        profile["check_batch"] = False

    # description witness: fit one TF-IDF vectoriser on ALL descriptions. Runs
    # every time now (matchcore package OR built-in fallback), so the report
    # always shows the vectorisation engine at work.
    _descs = ([get(d, "Description") for d in design_recs_noL0] +
              [get(d, "Revision Name") for d in design_recs_noL0] +
              [get(b, "Description") for b in built_recs_noL0])
    vectorizer = mc_build_vec(_descs)
    log(f"  Match engines: Material number + Description vectorisation "
        f"(TF-IDF, {_WITNESS_SOURCE}).")

    # What do we CALL the key on the built side? Adriele: a scan/SAP/mb51/ABCL
    # list has no "Part Number" — the key is the MATERIAL. Use that word in the
    # report whenever the built source is an adapted list; keep "Part Number"
    # only for a genuine Teamcenter As-Built export.
    key_label = ("Material" if any(b.get("_is_scan") for b in built_recs_noL0)
                 else "Part Number")

    # Auditable distinct-part reconciliation (headline numbers) written FIRST,
    # so the Summary + Parts Reconciliation are the opening tabs. The detailed
    # position/serial tabs follow.
    recon_rows, recon = part_reconciliation(design_recs, built_recs,
                                            label=built_label)
    if recon["built_parts"] == 0 and built_recs_noL0:
        log(f"  ! WARNING: read {len(built_recs_noL0)} rows from the As-Built "
            f"source but found NO part numbers in them. Check you picked the "
            f"right file/column (expected Part Number / Material / ID).")
    if recon["design_parts"] == 0 and design_recs_noL0:
        log("  ! WARNING: the As-Design has NO part numbers (expected an ID / "
            "Material column). Did you swap the As-Design and As-Built inputs?")
    write_summary_tab(wb, recon, name, label=built_label, meta=meta)
    write_reconciliation_tab(wb, recon_rows, key_label=key_label)

    deviations = write_matched_tab(wb, matched, mbom_cols_concise,
                                   asb_cols_concise,
                                   design_qty_by_id, built_qty_by_pn,
                                   profile=profile, vectorizer=vectorizer)
    # Column sets kept CLEAN so the tabs can be fed straight back into the next
    # round: the As-Built (extras) tab uses the canonical As-Built columns; the
    # As-Design (missing) tab — which becomes the NEXT As-Design — uses the concise
    # design set. Nothing else is carried through.
    built_is_list = any(b.get("_is_scan") for b in built_recs_noL0)
    extra_cols = (list(ASBUILT_CANONICAL) if built_is_list
                  else actual_cols(built_headers, built_recs_noL0))
    write_single_side_tab(
        wb, _sheet_safe(f"In {built_label}, not in Design"), unmatched_built,
        extra_cols,
        note=f"{recon['extra']} distinct part(s) IN {built_label.upper()} but "
             f"NOT IN THE DESIGN ({len(unmatched_built)} line(s); {key_label} "
             f"not in the design baseline)",
        note_fill=COL_DEVIATED)
    write_single_side_tab(
        wb, _sheet_safe(f"In Design, not in {built_label}"), unmatched_design,
        mbom_cols_concise,
        note=f"{recon['missing']} distinct part(s) IN THE DESIGN but NOT IN "
             f"{built_label.upper()} - shown here as {len(unmatched_design)} "
             f"BOM position(s) (same part can occupy several positions). See Summary.",
        note_fill=COL_MISSING)

    # THE DESCRIPTION ENGINE, made visible: rescue same-description/diff-number
    # pairs from the extras vs the missing. This is where the vectorisation earns
    # its keep — turning false extras/missing into reviewable real matches.
    rescued = write_description_rescue_tab(wb, unmatched_built, unmatched_design,
                                           vectorizer, key_label=key_label)
    log(f"  Name engine: {rescued} same-name candidate(s) found among the "
        f"{recon['extra']} extra / {recon['missing']} missing "
        f"(see 'Review - Name match').")

    try:
        wb.save(out_path)
    except PermissionError:
        log(f"  COULD NOT SAVE {os.path.basename(out_path)} — that report "
            f"is open in Excel. Close it and run again.")
        return {"unit": name, "error": "report open in Excel"}
    log(f"  {name}: DISTINCT parts -> matched {recon['matched']} "
        f"(short {recon['short']}, over {recon['over']}), "
        f"missing {recon['missing']}, extra {recon['extra']}  |  "
        f"positions: matched-lines {len(matched)}, "
        f"unmatched-built-lines {len(unmatched_built)}, "
        f"unmatched-design-positions {len(unmatched_design)}")
    return {"unit": name, "matched": len(matched), "deviations": deviations,
            "unmatched_built": len(unmatched_built),
            "unmatched_design": len(unmatched_design),
            "parts_matched": recon["matched"], "parts_missing": recon["missing"],
            "parts_extra": recon["extra"], "parts_short": recon["short"],
            "parts_over": recon["over"]}


def run_compare(design_arg, built_dir, out_dir, prefix="ADAB_Report",
                combine=False, progress=None, built_label="As-Built"):
    """Core comparison. Callable from CLI or GUI.

    design_arg  : As-Design — a single file OR a folder holding exactly one.
    built_dir   : As-Built — a single FILE or a FOLDER of .xls*/.xlsm files.
    out_dir     : folder where reports are written.
    prefix      : output filename prefix.
    combine     : if True, merge ALL As-Built files into one list and write a
                  single combined report — for when the As-Built files together
                  form ONE list/assembly. Default False = one report per file.
    progress    : optional callable(str) to report progress lines to a GUI.
    built_label : name of the As-Built source (Scan / Manual / Reserved / mb51 /
                  Teamcenter). Drives the contextual unmatched-tab names
                  ("In Design, not in <label>" / "In <label>, not in Design").

    Returns a list of result dicts. Overwrites existing reports.
    """
    def log(msg):
        if progress:
            progress(msg)
        else:
            print(msg)

    design_file = resolve_design(design_arg)
    design_base = os.path.splitext(os.path.basename(design_file))[0]
    try:
        design_headers, design_recs = load_bom(design_file)
    except PermissionError:
        raise ValueError(
            f"Cannot open the As-Design file:\n  {design_file}\n"
            "It looks like it's open in Excel — close it and run again.")
    except Exception as e:
        raise ValueError(
            f"Could not read the As-Design file:\n  {design_file}\n"
            f"{type(e).__name__}: {e}\n"
            "If it's on OneDrive, right-click it > 'Always keep on this device'.")
    design_recs_noL0 = [d for d in design_recs if not _is_level0(d)]
    mbom_cols_concise = present_nonempty(MBOM_COLS_CONCISE, design_headers,
                                         design_recs_noL0)
    mbom_cols_full = present_nonempty(MBOM_COLS_ALL, design_headers,
                                      design_recs_noL0)
    design_qty_by_id = {}
    for d in design_recs_noL0:
        k = design_key(d)
        q = _to_float(d.get("Quantity"))
        if k and q is not None:
            design_qty_by_id[k] = design_qty_by_id.get(k, 0.0) + q
    log(f"As-Designed: {os.path.basename(design_file)} "
        f"({len(design_recs_noL0)} parts, {len(mbom_cols_concise)} concise cols)")

    # As-Built input may be a single FILE or a FOLDER of files.
    if os.path.isfile(built_dir):
        built_files = [built_dir]
    else:
        built_files = sorted(
            f for f in glob.glob(os.path.join(built_dir, "*.xls*"))
            if not os.path.basename(f).startswith("~$")
        )
    if not built_files:
        raise ValueError("No As-Built file(s) found — pick a file or a folder "
                         "that contains .xlsx/.xlsm files.")

    os.makedirs(out_dir, exist_ok=True)
    results = []
    lead = prefix if prefix and prefix != "ADAB_Report" else "adab"

    # ---- COMBINE: merge every As-Built file into ONE list -> ONE report -----
    if combine:
        all_headers, all_recs, used = [], [], []
        for bf in built_files:
            try:
                bh, br = load_bom(bf)
            except Exception as e:
                log(f"  SKIPPED {os.path.basename(bf)}: "
                    f"{type(e).__name__}: {e}")
                continue
            all_recs.extend(br)
            for h in bh:
                if h not in all_headers:
                    all_headers.append(h)
            used.append(os.path.basename(bf))
        log(f"COMBINE: merged {len(used)} As-Built file(s) into one list "
            f"({len(all_recs)} rows) -> single 'full assembly' report.")
        out_path = os.path.join(
            out_dir, _safe_name(f"{lead}_{design_base}_ALL") + ".xlsx")
        results.append(_one_report(
            out_path, "ALL (full assembly)", all_headers, all_recs,
            design_recs, design_headers, design_recs_noL0,
            mbom_cols_concise, design_qty_by_id, log, built_label=built_label,
            design_name=os.path.basename(design_file)))
        log("\nDone. One combined report written (6 tabs).")
        return results

    # ---- default: one report per unit ---------------------------------------
    seen_names = {}
    for bf in built_files:
        name = sheet_name_for(bf)
        if name in seen_names:
            suffix = 2
            base = name
            while f"{base}_{suffix}" in seen_names:
                suffix += 1
            log(f"  WARNING: duplicate unit '{name}' "
                f"({os.path.basename(bf)}); writing as '{base}_{suffix}'")
            name = f"{base}_{suffix}"
        seen_names[name] = os.path.basename(bf)

        try:
            built_headers, built_recs = load_bom(bf)
        except PermissionError:
            log(f"  SKIPPED {name}: '{os.path.basename(bf)}' is open in Excel "
                f"(or locked). Close it and run again.")
            results.append({"unit": name, "error": "file open in Excel"})
            continue
        except Exception as e:
            log(f"  SKIPPED {name}: could not read '{os.path.basename(bf)}' "
                f"-> {type(e).__name__}: {e}")
            log("           (if it's on OneDrive, right-click > 'Always keep "
                "on this device')")
            results.append({"unit": name, "error": str(e)})
            continue

        built_base = os.path.splitext(os.path.basename(bf))[0]
        out_path = os.path.join(
            out_dir, _safe_name(f"{lead}_{design_base}_{built_base}") + ".xlsx")
        results.append(_one_report(
            out_path, name, built_headers, built_recs,
            design_recs, design_headers, design_recs_noL0,
            mbom_cols_concise, design_qty_by_id, log, built_label=built_label,
            design_name=os.path.basename(design_file)))

    log("\nDone. One report per unit written (6 tabs each).")
    return results


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--design", required=True, help="As-Designed file OR folder with one file")
    ap.add_argument("--built", required=True, help="Folder of As-Built .xlsm files")
    ap.add_argument("--out", default=".", help="Output folder for reports")
    ap.add_argument("--prefix", default="ADAB_Report",
                    help="Leading word of the output name. Default writes "
                         "adab_<as-design>_<as-built>.xlsx; a custom prefix "
                         "replaces the 'adab' lead.")
    ap.add_argument("--combine", action="store_true",
                    help="Merge ALL As-Built files into one and write a single "
                         "'full assembly' report (adab_<design>_ALL.xlsx), "
                         "instead of one report per unit.")
    args = ap.parse_args()
    run_compare(args.design, args.built, args.out, args.prefix,
                combine=args.combine)


if __name__ == "__main__":
    main()