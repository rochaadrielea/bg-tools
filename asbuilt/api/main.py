"""As-Built Tracker API.

Checkpoints are rows in `station`, not code — adding or retiring one is an API
call, never a rebuild. Migrations are numbered, applied once, and additive:
a deploy never drops a column, so the previous image still runs against the new
schema. That is what makes rollback a tag change.
"""
from __future__ import annotations

import io
import os
import pathlib
import datetime as dt

import openpyxl
import psycopg
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from psycopg.rows import dict_row
from pydantic import BaseModel

import engine as E

DSN = os.environ["DATABASE_URL"]
WEB = os.environ.get("WEB_DIR", "/app/web")
VERSION = os.environ.get("APP_VERSION", "dev")
MIGRATIONS = pathlib.Path(__file__).with_name("migrations")

app = FastAPI(title="As-Built Tracker")


def db():
    return psycopg.connect(DSN, row_factory=dict_row)


# --------------------------------------------------------------------------
# migrations — numbered, applied once, additive only
# --------------------------------------------------------------------------
def migrate() -> list:
    done = []
    with db() as c, c.cursor() as cur:
        cur.execute("""CREATE TABLE IF NOT EXISTS schema_migrations (
                         version TEXT PRIMARY KEY, applied_at TIMESTAMPTZ DEFAULT now())""")
        c.commit()
        cur.execute("SELECT version FROM schema_migrations")
        applied = {r["version"] for r in cur.fetchall()}
        for f in sorted(MIGRATIONS.glob("*.sql")):
            if f.stem in applied:
                continue
            cur.execute(f.read_text())
            cur.execute("INSERT INTO schema_migrations (version) VALUES (%s)", (f.stem,))
            c.commit()
            done.append(f.stem)
    return done


@app.on_event("startup")
def on_start():
    applied = migrate()
    if applied:
        print("applied migrations:", ", ".join(applied), flush=True)


# --------------------------------------------------------------------------
# reading
# --------------------------------------------------------------------------
def get_unit(cur, unit_id: int | None):
    if unit_id:
        cur.execute("SELECT * FROM unit WHERE id=%s", (unit_id,))
    else:
        cur.execute("SELECT * FROM unit ORDER BY id LIMIT 1")
    return cur.fetchone()


def stations(cur) -> list:
    cur.execute("SELECT * FROM station WHERE active ORDER BY position, id")
    return cur.fetchall()


def gates(cur) -> list:
    cur.execute("SELECT * FROM gate WHERE active ORDER BY position, id")
    return cur.fetchall()


def side_rows(cur, unit_id: int, ref: str) -> list | None:
    """A gate side: 'station:<key>' or 'bom:<kind>'. None = nothing uploaded."""
    kind, _, name = ref.partition(":")
    if kind == "bom":
        cur.execute("""SELECT material, revision, description, qty, traceable,
                              position, '' AS batch, '' AS serial
                       FROM bom WHERE unit_id=%s AND kind=%s""", (unit_id, name))
        rows = cur.fetchall()
        return rows or None
    cur.execute("""SELECT e.material, e.revision, e.description, e.qty, e.batch,
                          e.serial, e.position, FALSE AS traceable
                   FROM entry e JOIN station s ON s.id=e.station_id
                   WHERE e.unit_id=%s AND s.key=%s""", (unit_id, name))
    rows = cur.fetchall()
    return rows or None


def traceable_map(cur, unit_id: int) -> dict:
    """Traceable lives on the MBOM. It decides whether a missing batch is a
    flag, so it has to follow the material into every station's rows."""
    cur.execute("SELECT material, traceable FROM bom WHERE unit_id=%s AND kind='mbom'",
                (unit_id,))
    return {E.match_key(r["material"]): r["traceable"] for r in cur.fetchall()}


def apply_traceable(rows: list, tmap: dict) -> list:
    for r in rows:
        if not r.get("traceable"):
            r["traceable"] = tmap.get(E.match_key(r["material"]), False)
    return rows


@app.get("/api/state")
def state(unit_id: int | None = None):
    with db() as c, c.cursor() as cur:
        u = get_unit(cur, unit_id)
        cur.execute("SELECT id, equipment_no, serial, product FROM unit ORDER BY id")
        units = cur.fetchall()
        if not u:
            return {"version": VERSION, "unit": None, "units": [],
                    "stations": stations(cur), "gates": gates(cur), "chain": []}

        sts, gts = stations(cur), gates(cur)
        cur.execute("""SELECT DISTINCT ON (station_id) station_id, filename, uploaded_at,
                              rows_loaded
                       FROM upload WHERE unit_id=%s
                       ORDER BY station_id, uploaded_at DESC""", (u["id"],))
        ups = {r["station_id"]: r for r in cur.fetchall()}
        for s in sts:
            up = ups.get(s["id"])
            s["uploaded"] = bool(up)
            s["filename"] = up["filename"] if up else None
            s["uploaded_at"] = up["uploaded_at"].strftime("%d/%m/%Y") if up else None
            s["rows_loaded"] = up["rows_loaded"] if up else 0

        tmap = traceable_map(cur, u["id"])
        for g in gts:
            a = side_rows(cur, u["id"], g["expected_ref"])
            b = side_rows(cur, u["id"], g["present_ref"])
            g["ready"] = a is not None and b is not None
            if g["ready"]:
                res = E.compare(apply_traceable(a, tmap), apply_traceable(b, tmap))
                g["counts"] = res["counts"]
            else:
                g["counts"] = None

        frontier = None
        for s in sts:
            if s["uploaded"]:
                frontier = s
        return {"version": VERSION, "unit": u, "units": units,
                "stations": sts, "gates": gts, "frontier": frontier}


@app.get("/api/gate/{gate_key}")
def gate_detail(gate_key: str, unit_id: int | None = None):
    with db() as c, c.cursor() as cur:
        u = get_unit(cur, unit_id)
        if not u:
            raise HTTPException(404, "no unit")
        cur.execute("SELECT * FROM gate WHERE key=%s", (gate_key,))
        g = cur.fetchone()
        if not g:
            raise HTTPException(404, "no such gate")
        a = side_rows(cur, u["id"], g["expected_ref"])
        b = side_rows(cur, u["id"], g["present_ref"])
        if a is None or b is None:
            missing = g["label_a"] if a is None else g["label_b"]
            raise HTTPException(409, f"nothing uploaded yet for {missing}")
        tmap = traceable_map(cur, u["id"])
        res = E.compare(apply_traceable(a, tmap), apply_traceable(b, tmap))
    res["gate"] = g
    res["labels"] = E.VERDICT_LABEL
    return res


# --------------------------------------------------------------------------
# uploading — the file is stored byte for byte, so an import can be replayed
# --------------------------------------------------------------------------
def sheet_rows(content: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best = []
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        if len(rows) > len(best):
            best = rows
    return best


@app.post("/api/upload/{station_key}")
async def upload(station_key: str, file: UploadFile = File(...),
                 unit_id: int | None = Form(None)):
    content = await file.read()
    try:
        rows, fields = E.read_rows(sheet_rows(content))
    except ValueError as e:
        raise HTTPException(400, str(e))

    with db() as c, c.cursor() as cur:
        u = get_unit(cur, unit_id)
        if not u:
            raise HTTPException(404, "no unit")
        cur.execute("SELECT id FROM station WHERE key=%s AND active", (station_key,))
        s = cur.fetchone()
        if not s:
            raise HTTPException(404, f"no station '{station_key}'")

        # a station holds one truth: the latest file for it replaces the last
        cur.execute("DELETE FROM entry WHERE unit_id=%s AND station_id=%s", (u["id"], s["id"]))
        for r in rows:
            cur.execute("""INSERT INTO entry (unit_id, station_id, material, revision,
                              description, batch, serial, qty, work_order, position, source_file)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (u["id"], s["id"], r.get("material", ""), r.get("revision", ""),
                         r.get("description", ""), r.get("batch", ""), r.get("serial", ""),
                         r.get("qty", 0), r.get("work_order", ""),
                         int(E.to_num(r.get("position", 1)) or 1), file.filename))
        cur.execute("""INSERT INTO upload (unit_id, station_id, filename, rows_loaded, content)
                       VALUES (%s,%s,%s,%s,%s)""",
                    (u["id"], s["id"], file.filename, len(rows), content))
        c.commit()
    return {"station": station_key, "rows": len(rows), "fields": fields,
            "filename": file.filename}


@app.delete("/api/upload/{station_key}")
def clear_station(station_key: str, unit_id: int | None = None):
    """Remove a station's data. The uploaded files stay, so it can be replayed."""
    with db() as c, c.cursor() as cur:
        u = get_unit(cur, unit_id)
        cur.execute("SELECT id FROM station WHERE key=%s", (station_key,))
        s = cur.fetchone()
        if not (u and s):
            raise HTTPException(404, "no such unit or station")
        cur.execute("DELETE FROM entry WHERE unit_id=%s AND station_id=%s", (u["id"], s["id"]))
        cur.execute("DELETE FROM upload WHERE unit_id=%s AND station_id=%s", (u["id"], s["id"]))
        c.commit()
    return {"cleared": station_key}


@app.get("/api/upload/{station_key}/file")
def upload_file(station_key: str, unit_id: int | None = None):
    with db() as c, c.cursor() as cur:
        u = get_unit(cur, unit_id)
        cur.execute("""SELECT up.filename, up.content FROM upload up
                       JOIN station s ON s.id=up.station_id
                       WHERE up.unit_id=%s AND s.key=%s
                       ORDER BY up.uploaded_at DESC LIMIT 1""", (u["id"], station_key))
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "nothing uploaded for this station yet")
    return Response(row["content"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{row["filename"]}"'})


# --------------------------------------------------------------------------
# checkpoints are data: add, edit, retire — no rebuild
# --------------------------------------------------------------------------

class UnitIn(BaseModel):
    equipment_no: str
    serial: str = ""
    order_no: str = ""
    sach_nr: str = ""
    product: str = ""


@app.post("/api/unit")
def add_unit(body: UnitIn):
    """A new unit — needed before any station can hold uploaded rows."""
    if not body.equipment_no.strip():
        raise HTTPException(400, "equipment_no is required")
    with db() as c, c.cursor() as cur:
        cur.execute("""INSERT INTO unit (equipment_no, serial, order_no, sach_nr, product)
                       VALUES (%s,%s,%s,%s,%s) RETURNING *""",
                    (body.equipment_no.strip(), body.serial.strip(),
                     body.order_no.strip(), body.sach_nr.strip(), body.product.strip()))
        row = cur.fetchone()
        c.commit()
    return row


class StationIn(BaseModel):
    key: str
    name: str
    source: str = ""
    icon: str = "•"
    position: int | None = None
    after: str | None = None        # insert straight after this station's key


class GateIn(BaseModel):
    key: str
    expected_ref: str
    present_ref: str
    label_a: str = ""
    label_b: str = ""
    position: int | None = None


@app.post("/api/station")
def add_station(body: StationIn):
    with db() as c, c.cursor() as cur:
        pos = body.position
        if pos is None and body.after:
            cur.execute("SELECT position FROM station WHERE key=%s", (body.after,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(404, f"no station '{body.after}' to insert after")
            pos = r["position"] + 1
        if pos is None:
            cur.execute("SELECT COALESCE(MAX(position),0)+1 AS p FROM station")
            pos = cur.fetchone()["p"]
        cur.execute("UPDATE station SET position=position+1 WHERE position>=%s", (pos,))
        cur.execute("""INSERT INTO station (key,name,source,icon,position)
                       VALUES (%s,%s,%s,%s,%s) RETURNING *""",
                    (body.key, body.name, body.source, body.icon, pos))
        row = cur.fetchone()
        c.commit()
    return row


@app.delete("/api/station/{key}")
def retire_station(key: str, hard: bool = False):
    """Default is a soft retire — the rows and files stay, so it is reversible.
    hard=true deletes the station and everything under it."""
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT id FROM station WHERE key=%s", (key,))
        if not cur.fetchone():
            raise HTTPException(404, f"no station '{key}'")
        if hard:
            cur.execute("DELETE FROM station WHERE key=%s", (key,))
        else:
            cur.execute("UPDATE station SET active=FALSE WHERE key=%s", (key,))
        cur.execute("""UPDATE gate SET active=FALSE
                       WHERE expected_ref=%s OR present_ref=%s""",
                    (f"station:{key}", f"station:{key}"))
        c.commit()
    return {"retired": key, "hard": hard}


@app.post("/api/station/{key}/restore")
def restore_station(key: str):
    with db() as c, c.cursor() as cur:
        cur.execute("UPDATE station SET active=TRUE WHERE key=%s RETURNING *", (key,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, f"no station '{key}'")
        c.commit()
    return row


@app.post("/api/gate")
def add_gate(body: GateIn):
    with db() as c, c.cursor() as cur:
        pos = body.position
        if pos is None:
            cur.execute("SELECT COALESCE(MAX(position),0)+1 AS p FROM gate")
            pos = cur.fetchone()["p"]
        cur.execute("""INSERT INTO gate (key,expected_ref,present_ref,label_a,label_b,position)
                       VALUES (%s,%s,%s,%s,%s,%s) RETURNING *""",
                    (body.key, body.expected_ref, body.present_ref,
                     body.label_a, body.label_b, pos))
        row = cur.fetchone()
        c.commit()
    return row


@app.delete("/api/gate/{key}")
def retire_gate(key: str, hard: bool = False):
    with db() as c, c.cursor() as cur:
        if hard:
            cur.execute("DELETE FROM gate WHERE key=%s", (key,))
        else:
            cur.execute("UPDATE gate SET active=FALSE WHERE key=%s", (key,))
        c.commit()
    return {"retired": key, "hard": hard}


# --------------------------------------------------------------------------
# export
# --------------------------------------------------------------------------

@app.get("/api/trace")
def trace(material: str = "", batch: str = "", serial: str = "",
          project: str = "", date_from: str = "", date_to: str = ""):
    """Trace a material / batch / serial across the chain.

    Defaults for scope:
      - material search -> current unit (project= names it; empty = every unit)
      - batch / serial search -> every unit unless project= is given

    Missing station is a fact shown as a grey cell, not an error. A traceable
    part with no batch recorded stays flagged in red — that's the point.
    """
    if not (material or batch or serial):
        raise HTTPException(400, "give at least one of material, batch or serial")

    mk = E.match_key(material) if material else ""
    bk = E.trace_key(batch)   if batch    else ""
    sk = E.trace_key(serial)  if serial   else ""

    where, args = [], []
    if mk:
        # material joins are type-blind and suffix-preserving — do the compare
        # in the database using the same rule (UPPER on a stripped copy)
        where.append("UPPER(REGEXP_REPLACE(e.material,'\\s','','g')) = %s")
        args.append(mk)
    if bk:
        where.append("UPPER(REGEXP_REPLACE(e.batch,'[\\s\\-.\\/_]','','g')) = %s")
        args.append(bk)
    if sk:
        where.append("UPPER(REGEXP_REPLACE(e.serial,'[\\s\\-.\\/_]','','g')) = %s")
        args.append(sk)
    if project:
        where.append("u.equipment_no = %s")
        args.append(project)
    if date_from:
        where.append("e.imported_at >= %s")
        args.append(date_from)
    if date_to:
        where.append("e.imported_at <= %s")
        args.append(date_to + " 23:59:59")

    with db() as c, c.cursor() as cur:
        cur.execute(f"""
            SELECT u.id AS unit_id, u.equipment_no, u.serial AS unit_serial,
                   s.key AS station_key, s.name AS station_name, s.position,
                   e.material, e.revision, e.description, e.batch, e.serial,
                   e.qty, e.work_order, e.imported_at, e.source_file
            FROM entry e
            JOIN unit u    ON u.id = e.unit_id
            JOIN station s ON s.id = e.station_id
            WHERE {' AND '.join(where)}
            ORDER BY u.equipment_no, s.position, e.imported_at
        """, args)
        entry_rows = cur.fetchall()

        # BOM says whether the part is expected to carry a batch (Traceable)
        cur.execute(f"""
            SELECT u.id AS unit_id, u.equipment_no, b.material, b.revision,
                   b.description, b.qty, b.traceable, b.kind
            FROM bom b JOIN unit u ON u.id = b.unit_id
            WHERE b.kind IN ('mbom','ebom')
              {"AND UPPER(REGEXP_REPLACE(b.material,'\\s','','g')) = %s" if mk else ""}
              {"AND u.equipment_no = %s" if project else ""}
        """, [x for x in [mk, project] if x])
        bom_rows = cur.fetchall()

        cur.execute("""SELECT DISTINCT ON (s.id) s.key, s.name, s.position
                       FROM station s WHERE s.active ORDER BY s.id, s.position""")
        stations = [{"key":r["key"], "name":r["name"], "position":r["position"]}
                    for r in cur.fetchall()]
        stations.sort(key=lambda x: x["position"])

    # traceable flag per (unit, material) key from the MBOM
    tmap = {}
    for r in bom_rows:
        if r["kind"] == "mbom":
            tmap[(r["unit_id"], E.match_key(r["material"]))] = bool(r["traceable"])

    # group by unit -> station
    units = {}
    for r in entry_rows:
        u = units.setdefault(r["equipment_no"], {
            "equipment_no": r["equipment_no"],
            "unit_serial": r["unit_serial"],
            "by_station": {},
        })
        st = u["by_station"].setdefault(r["station_key"], {
            "name": r["station_name"], "position": r["position"],
            "rows": [], "qty": 0.0, "batches": set(), "serials": set(),
        })
        st["rows"].append({
            "material": r["material"], "revision": r["revision"],
            "description": r["description"], "batch": r["batch"],
            "serial": r["serial"], "qty": float(r["qty"] or 0),
            "work_order": r["work_order"],
            "imported_at": r["imported_at"].isoformat() if r["imported_at"] else "",
            "source_file": r["source_file"],
        })
        st["qty"] += float(r["qty"] or 0)
        if r["batch"]:  st["batches"].add(r["batch"])
        if r["serial"]: st["serials"].add(r["serial"])

    # expected-side rows (MBOM/EBOM) so unit cards appear even when nothing has
    # been received yet — that's the "did the project want this part" answer
    for r in bom_rows:
        eq = r["equipment_no"]
        u = units.setdefault(eq, {
            "equipment_no": eq, "unit_serial": "",
            "by_station": {}, "expected": []
        })
        u.setdefault("expected", []).append({
            "kind": r["kind"], "material": r["material"],
            "revision": r["revision"], "description": r["description"],
            "qty": float(r["qty"] or 0), "traceable": bool(r["traceable"]),
        })

    # shape into a stable list for the front end
    out_units = []
    for eq in sorted(units):
        u = units[eq]
        chain = []
        for s in stations:
            hit = u["by_station"].get(s["key"])
            if hit:
                chain.append({
                    "station": s["key"], "name": s["name"],
                    "present": True, "qty": hit["qty"],
                    "batches": sorted(hit["batches"]),
                    "serials": sorted(hit["serials"]),
                    "rows": hit["rows"],
                    "flag": (mk and tmap.get((next((x["unit_id"] for x in bom_rows
                                                    if x["equipment_no"]==eq), None), mk))
                            and not hit["batches"]),
                })
            else:
                chain.append({"station": s["key"], "name": s["name"],
                              "present": False})
        out_units.append({
            "equipment_no": eq, "unit_serial": u.get("unit_serial",""),
            "chain": chain, "expected": u.get("expected", []),
        })

    return {
        "query": {"material": material, "batch": batch, "serial": serial,
                  "project": project, "date_from": date_from, "date_to": date_to,
                  "material_key": mk, "batch_key": bk, "serial_key": sk},
        "units": out_units,
        "stations": stations,
        "counts": {"units": len(out_units),
                   "hits": sum(1 for u in out_units
                               for s in u["chain"] if s["present"])},
    }


@app.get("/api/projects")
def projects():
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT equipment_no, serial FROM unit ORDER BY equipment_no")
        return {"projects": cur.fetchall()}



@app.get("/api/export/{gate_key}")
def export_gate(gate_key: str, unit_id: int | None = None):
    res = gate_detail(gate_key, unit_id)
    g = res["gate"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = gate_key[:31]
    ws.append([f"{g['label_a']} (expected)", "", "", "", "",
               f"{g['label_b']} (present)", "", "", "", "", ""])
    ws.append(["Material", "Description", "Revision", "Qty", "Positions",
               "Revision", "Qty", "Batches", "Serials", "Verdict", "Detail"])
    for r in res["rows"]:
        a, b = r["a"], r["b"]
        ws.append([
            (a or b)["material"], (a or b).get("description", ""),
            "/".join(a["revisions"]) if a else "", E.fmt_qty(a["qty"]) if a else "",
            a["positions"] if a else "",
            "/".join(b["revisions"]) if b else "", E.fmt_qty(b["qty"]) if b else "",
            ", ".join(b["batches"]) if b else "", ", ".join(b["serials"]) if b else "",
            E.VERDICT_LABEL[r["verdict"]], r["detail"]])
    buf = io.BytesIO()
    wb.save(buf)
    name = f"AsBuilt_{gate_key}_{dt.date.today():%Y-%m-%d}.xlsx"
    return Response(buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'})


# --------------------------------------------------------------------------
@app.get("/health")
def health():
    with db() as c, c.cursor() as cur:
        cur.execute("SELECT count(*) AS n FROM station WHERE active")
        n = cur.fetchone()["n"]
        cur.execute("SELECT count(*) AS n FROM entry")
        e = cur.fetchone()["n"]
        cur.execute("SELECT version FROM schema_migrations ORDER BY version DESC LIMIT 1")
        m = cur.fetchone()
    return {"ok": True, "version": VERSION, "stations": n, "entries": e,
            "schema": m["version"] if m else None}


@app.get("/")
def index():
    return FileResponse(os.path.join(WEB, "index.html"))


@app.exception_handler(HTTPException)
def http_error(_, exc: HTTPException):
    return JSONResponse({"error": exc.detail}, status_code=exc.status_code)